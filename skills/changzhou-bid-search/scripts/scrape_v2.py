#!/usr/bin/env python3
"""常州招标抓取脚本 v2 - Python版（2026-04 网站改版适配）"""
import urllib.request, re, ssl, time, os, io, zipfile, sys, json, csv, html as html_mod, shutil, subprocess
from datetime import datetime, timedelta
from urllib.parse import urlencode
from pathlib import Path

BASE = 'http://ggzy.xzsp.changzhou.gov.cn'
SITE_GUID = '7eb5f7f1-9041-43ad-8e13-8fcb82ea831a'
HEADERS = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}

# ── 腾讯文档同步配置 ────────────────────────────────────────────────────────
TENCENT_FILE_ID = 'XvrbguxMdXjt'   # 腾讯文档表格 ID
TENCENT_SHEET_ID = 'BB08J2'        # 腾讯文档子表 ID
TENCENT_URL = 'https://docs.qq.com/sheet/DWHZyYmd1eE1kWGp0'
OUT_DIR = '/workspace/skills/changzhou-bid-search/outputs'
MAX_DAYS = 7                       # 保留近 N 天内的公告
PERSISTENT_EXCEL = OUT_DIR + '/常州招标_持续积累.xlsx'

# ── 日期工具 ────────────────────────────────────────────────────────────────
def parse_date(s):
    """将 YYYY-MM-DD 字符串解析为 date 对象"""
    for fmt in ('%Y-%m-%d', '%Y/%m/%d'):
        try:
            return datetime.strptime(s.strip()[:10], fmt).date()
        except ValueError:
            pass
    return None

def is_within_days(date_str, days=MAX_DAYS):
    """判断日期是否在近 N 天内"""
    d = parse_date(date_str)
    if d is None:
        return False
    cutoff = datetime.now().date() - timedelta(days=days)
    return d >= cutoff

def is_expired(date_str, days=MAX_DAYS):
    """判断日期是否已过期（超过 N 天）"""
    d = parse_date(date_str)
    if d is None:
        return False
    cutoff = datetime.now().date() - timedelta(days=days)
    return d < cutoff

# ── 腾讯文档读取（返回 rows list）────────────────────────────────────────────
def qq_read_all(mcporter_cmd='mcporter'):
    """
    从腾讯文档读取当前全部数据（section_id + date + title）。
    返回: [{'row_idx': 1-based行号, 'section_id': '', 'date': '', 'title': ''}, ...]
    空列表表示文档为空或 API 问题。
    mcporter stdout 有 64KB 限制，正常使用（≤200行）下不会触发。
    """
    import io as sio

    cmd = [
        mcporter_cmd, 'call', 'tencent-sheetengine', 'get_cell_data',
        '--args', json.dumps({
            'file_id': TENCENT_FILE_ID,
            'sheet_id': TENCENT_SHEET_ID,
            'start_row': 0,
            'start_col': 0,
            'end_row': 2000,  # 2000行 ≈ 200KB（接近 mcporter stdout 64KB 上限，但够用）
            'end_col': 23,
            'return_csv': True,
        })
    ]
    try:
        result = subprocess.run(
            cmd, capture_output=True, timeout=30,
            text=False   # 返回 bytes
        )
        raw = result.stdout
    except Exception as e:
        print(f'    [WARN] 腾讯文档读取失败: {e}')
        return []

    # 尝试解码，遇到编码问题截断
    try:
        text = raw.decode('utf-8')
    except Exception:
        text = raw.decode('utf-8', errors='replace')

    try:
        resp = json.loads(text)
    except Exception as e:
        # JSON 不完整（mcporter stdout 64KB 截断）
        # 仅当已有历史数据时才应关注此问题
        print(f'    [WARN] 腾讯文档响应解析失败（{len(text)} bytes），可能数据被截断: {e}')
        return []

    csv_data = resp.get('csv_data', '')
    if not csv_data or not csv_data.strip():
        return []

    reader = csv.DictReader(sio.StringIO(csv_data))
    rows = []
    for i, row in enumerate(reader):
        sid = row.get('标段编号', '').strip()
        # 跳过空行
        if not sid:
            continue
        rows.append({
            'row_idx': i,           # 0-based（API 返回的 CSV row index = Excel row index）
            'section_id': sid,
            'date': row.get('发布日期', '').strip(),
            'title': row.get('项目名称', '').strip(),
        })

    return rows

# ── 腾讯文档增量写入 ────────────────────────────────────────────────────────
def qq_sync(mcporter_cmd, local_items, dry_run=False):
    """
    将 local_items 与腾讯文档同步：
      - 读取当前文档全部 section_id
      - 清理已过期项目（>7天，从文档+本地PDF删除）
      - 合并保留项目（文档中≤7天的 + 本次新增的）
      - 全量重写（clear_range_cells + set_range_value，避免行定位问题）
    返回: (新增数, 删除数)
    """
    print(f'\n[Step 4] 腾讯文档同步（保留近 {MAX_DAYS} 天）...')

    # ── Step A: 读取当前文档全部 section_id ───────────────────────────────
    existing = qq_read_all(mcporter_cmd)
    existing_map = {r['section_id']: r for r in existing if r.get('section_id')}

    # ── Step B: 分类：过期 vs 保留 vs 新增 ───────────────────────────────
    valid_items = [it for it in local_items
                   if it.get('date') and is_within_days(it['date'], MAX_DAYS)]

    # 已在文档中且未过期 → 保留
    kept_items = [it for it in valid_items
                  if it.get('section_id') in existing_map
                  and not is_expired(it.get('date', ''), MAX_DAYS)]

    # 本次新增（文档中无）→ 新增
    new_items = [it for it in valid_items
                 if it.get('section_id') not in existing_map]

    # 文档中有但本次抓取无（已过期）→ 需删除
    current_sids = {it.get('section_id') for it in valid_items if it.get('section_id')}
    expired_items = [r for sid, r in existing_map.items()
                      if sid not in current_sids and is_expired(r.get('date', ''), MAX_DAYS)]

    print(f'    文档现有: {len(existing_map)} 项 | 保留: {len(kept_items)} | 新增: {len(new_items)} | 过期删除: {len(expired_items)}')
    for it in new_items:
        print(f'      + {it.get("date")} | {it.get("title", it.get("project_name",""))[:40]}')
    for e in expired_items:
        print(f'      - {e.get("date")} | {e.get("section_id", "")[:30]} | {e.get("title","")[:30]}')

    if dry_run:
        print('    [DRY RUN] 跳过实际写入')
        return len(new_items), len(expired_items)

    # ── Step C: 清理过期本地 PDF 文件夹 ─────────────────────────────────
    for e in expired_items:
        sid = e.get('section_id', '')
        for fname in os.listdir(OUT_DIR):
            fpath = os.path.join(OUT_DIR, fname)
            if os.path.isdir(fpath) and (sid in fname or fname.endswith(sid[:8])):
                shutil.rmtree(fpath)
                print(f'    删除过期 PDF: {fname}')

    # ── Step D: 合并 + 全量重写 ─────────────────────────────────────────
    # 按日期倒序排列（最新的在前）
    all_items = sorted(
        valid_items,
        key=lambda x: x.get('date', ''),
        reverse=True
    )

    # 构建 cells: row 0 = header, rows 1~N = data
    headers = ['序号', '项目名称', '地区', '发布日期', '分类',
               '招标人', '招标代理', '建设地点', '建设内容',
               '控制价(万元)', '评标办法', '资格审查',
               '投标保证金', '工期', '计划开工', '计划竣工',
               '资质要求', '项目负责人资质', '合同价格形式', '投标有效期',
               '履约担保', '联系方式', '标段编号', '来源']

    cells = []
    for ci, h in enumerate(headers):
        cells.append({'row': 0, 'col': ci, 'value_type': 'STRING', 'string_value': h})

    for ri_offset, it in enumerate(all_items):
        row_num = ri_offset + 1   # 数据从 row 1 开始（row 0 是表头）
        vals = [
            str(ri_offset + 1),
            it.get('project_name', '') or it.get('title', ''),
            it.get('area', ''),
            it.get('date', ''),
            it.get('sub_cat', ''),
            it.get('client', ''),
            it.get('agent', ''),
            it.get('site', ''),
            (it.get('content', '') or '')[:300],   # 建设内容截断防超长
            it.get('budget', ''),
            it.get('bid_method', ''),
            it.get('qual_check', ''),
            it.get('deposit', ''),
            it.get('duration', ''),
            it.get('start_date', ''),
            it.get('end_date', ''),
            it.get('qual_req', ''),
            it.get('pm_req', ''),
            it.get('price_type', ''),
            it.get('valid_period', ''),
            it.get('perform_guarantee', ''),
            it.get('contact', '') or it.get('agent_contact', ''),
            it.get('section_id', ''),
            '常州公共资源交易中心',
        ]
        for ci, v in enumerate(vals):
            cells.append({'row': row_num, 'col': ci, 'value_type': 'STRING',
                           'string_value': str(v)[:200]})

    # ── Step E: 全量清空 + CSV 写入（CSV 格式更稳定）────────────────────
    last_row = max(len(all_items), len(existing_map)) + 10
    try:
        # 清空数据区域（start_row=1 跳过表头 row 0）
        subprocess.check_output([
            mcporter_cmd, 'call', 'tencent-sheetengine', 'clear_range_cells',
            '--args', json.dumps({
                'file_id': TENCENT_FILE_ID, 'sheet_id': TENCENT_SHEET_ID,
                'start_row': 1, 'start_col': 0, 'end_row': last_row, 'end_col': 23,
            })
        ], stderr=subprocess.DEVNULL, timeout=30)

        # 构建 CSV 内容（set_range_value_by_csv 更稳定，避免 JSON cells 截断问题）
        csv_out = io.StringIO()
        writer = csv.writer(csv_out)
        writer.writerow(headers)  # 表头 → row 0
        for ri_offset, it in enumerate(all_items):
            row_vals = [
                str(ri_offset + 1),
                it.get('project_name', '') or it.get('title', ''),
                it.get('area', ''),
                it.get('date', ''),
                it.get('sub_cat', ''),
                it.get('client', ''),
                it.get('agent', ''),
                it.get('site', ''),
                (it.get('content', '') or '')[:300],
                it.get('budget', ''),
                it.get('bid_method', ''),
                it.get('qual_check', ''),
                it.get('deposit', ''),
                it.get('duration', ''),
                it.get('start_date', ''),
                it.get('end_date', ''),
                it.get('qual_req', ''),
                it.get('pm_req', ''),
                it.get('price_type', ''),
                it.get('valid_period', ''),
                it.get('perform_guarantee', ''),
                it.get('contact', '') or it.get('agent_contact', ''),
                it.get('section_id', ''),
                '常州公共资源交易中心',
            ]
            writer.writerow(row_vals)

        csv_str = csv_out.getvalue()
        # 使用 set_range_value_by_csv 写入（比 set_range_value 更稳定）
        subprocess.check_output([
            mcporter_cmd, 'call', 'tencent-sheetengine', 'set_range_value_by_csv',
            '--args', json.dumps({
                'file_id': TENCENT_FILE_ID,
                'sheet_id': TENCENT_SHEET_ID,
                'csv_data': csv_str,
                'start_col': 0,
            })
        ], stderr=subprocess.DEVNULL, timeout=30)
        print(f'    全量写入 {len(all_items)} 行 ✅')

        # ── 验证写入 ─────────────────────────────────────────────────
        verify = qq_read_all(mcporter_cmd)
        written_count = len([r for r in verify if r.get('section_id')])
        if written_count != len(all_items):
            print(f'    [WARN] 写入验证失败：期望 {len(all_items)} 行，实际 {written_count} 行，重试...')
            subprocess.check_output([
                mcporter_cmd, 'call', 'tencent-sheetengine', 'set_range_value_by_csv',
                '--args', json.dumps({
                    'file_id': TENCENT_FILE_ID,
                    'sheet_id': TENCENT_SHEET_ID,
                    'csv_data': csv_str,
                    'start_col': 0,
                })
            ], stderr=subprocess.DEVNULL, timeout=30)
            verify2 = qq_read_all(mcporter_cmd)
            written2 = len([r for r in verify2 if r.get('section_id')])
            print(f'    重试后：{written2} 行')
        else:
            print(f'    验证通过：{written_count} 行')
    except Exception as ex:
        print(f'    腾讯文档写入失败: {ex}')

    return len(new_items), len(expired_items)

# ── HTTP 工具 ───────────────────────────────────────────────────────────────
def get(url, headers=None):
    ctx = ssl.create_default_context()
    ctx.check_hostname = False; ctx.verify_mode = ssl.CERT_NONE
    h = dict(HEADERS)
    if headers: h.update(headers)
    req = urllib.request.Request(url, headers=h)
    with urllib.request.urlopen(req, timeout=15, context=ctx) as r:
        return r.read().decode('utf-8', 'ignore')

def post(url, data, headers=None):
    ctx = ssl.create_default_context()
    ctx.check_hostname = False; ctx.verify_mode = ssl.CERT_NONE
    h = dict(HEADERS)
    if headers: h.update(headers)
    enc = urlencode(data).encode('utf-8') if isinstance(data, dict) else data
    req = urllib.request.Request(url, data=enc, headers=h)
    with urllib.request.urlopen(req, timeout=15, context=ctx) as r:
        return r.read().decode('utf-8', 'ignore')

def download(url, out_path):
    """下载文件到 out_path，返回文件大小（字节）"""
    ctx = ssl.create_default_context()
    ctx.check_hostname = False; ctx.verify_mode = ssl.CERT_NONE
    req = urllib.request.Request(url, headers=dict(HEADERS))
    with urllib.request.urlopen(req, timeout=30, context=ctx) as r:
        data = r.read()
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, 'wb') as f:
        f.write(data)
    return len(data)

# ── 常州列表页解析 ───────────────────────────────────────────────────────────
def parse_list(html):
    items, seen = [], set()
    for m in re.finditer(r'<tr[^>]*>([\s\S]*?)</tr>', html):
        row = m.group(1)
        segs = row.split('</td>')
        if len(segs) < 4:
            continue
        om = re.search(r"tzjydetail\s*\(\s*'([^']+)'\s*,\s*'([^']+)'\s*,", segs[1])
        if not om or om.group(2) in segs[0]:
            continue
        tm = re.search(r'title="([^"]{5,120})"', segs[1])
        if not tm:
            continue
        uuid = om.group(2)
        if uuid in seen:
            continue
        seen.add(uuid)
        area = re.sub(r'<[^>]+>', '', segs[2]).replace('\xa0', ' ').strip()
        date = re.sub(r'<[^>]+>', '', segs[3]).replace('\xa0', ' ').strip()
        items.append({
            'cat':   om.group(1),
            'uuid':  uuid,
            'title': tm.group(1)[:120],
            'area':  area,
            'date':  date,
        })
    return items

# ── 详情页 URL 获取（AJAX）──────────────────────────────────────────────────
def get_detail_url(uuid, cat):
    """通过 AJAX 获取详情页实际 URL"""
    try:
        resp = post(
            BASE + '/czggzyweb/frontPageRedirctAction.action?cmd=pageRedirect',
            {'infoid': uuid, 'siteGuid': SITE_GUID, 'categorynum': cat},
            headers={
                'X-Requested-With': 'XMLHttpRequest',
                'Accept': 'application/json',
                'Referer': BASE + '/jyzx/001001/tradeInfonew.html',
                'Content-Type': 'application/x-www-form-urlencoded',
            }
        )
        data = json.loads(resp)
        return BASE + data['custom'] if data.get('custom') else None
    except Exception as e:
        return None

# ── 正文文本清洗 ────────────────────────────────────────────────────────────
def clean_text(text):
    text = re.sub(r'<script[^>]*>[\s\S]*?</script>', ' ', text)
    text = re.sub(r'<style[^>]*>[\s\S]*?</style>', ' ', text)
    text = re.sub(r'<[^>]+>', ' ', text)
    text = html_mod.unescape(text)
    text = text.replace('\xa0', ' ').replace('&nbsp;', ' ')
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def find_field(text, keywords, length=300):
    for kw in keywords:
        i = text.find(kw)
        if i == -1:
            continue
        snippet = text[i:i+length]
        ci = snippet.find('：')
        ni = snippet.find('\n')
        end = min(ci + 1 if ci != -1 else length, ni if ni != -1 else length)
        val = snippet[ci+1:end].strip() if ci != -1 else snippet[:end].strip()
        if len(val) > 1:
            return val[:200]
    return ''

# ── 详情页字段提取 ───────────────────────────────────────────────────────────
def extract_table_fields(html):
    """
    从详情页的表格结构中提取键值对。
    支持两类结构：
    1. ewb-trade-right 中的嵌套 HTML 文档（中标结果公告等）
    2. 直接的 <td> 表格
    返回 dict
    """
    result = {}

    # 策略：找所有 <td> 的 align 属性和文本
    # 格式: <td align="center" bgcolor='#e5f2fa'>标签名</td><td align="left" ...>值</td>
    # 配对相邻的 "标签" td 和 "值" td
    td_pattern = re.compile(
        r'<td[^>]*(?:align|bgcolor)[^>]*>\s*([^<]{1,100})\s*</td>'
    )
    tds = td_pattern.findall(html)
    # 清理
    tds = [re.sub(r'<[^>]+>', '', t).replace('\xa0', ' ').strip() for t in tds]
    tds = [html_mod.unescape(t) for t in tds]
    tds = [re.sub(r'\s+', ' ', t) for t in tds]

    # 奇数位是 key，偶数位是 value
    for i in range(0, len(tds)-1, 2):
        key = tds[i].rstrip('：:').strip()
        val = tds[i+1].strip()
        if key and val:
            result[key] = val

    return result

def extract_detail_fields(html):
    """从详情页 HTML 提取投标相关字段"""
    # 先尝试表格提取（适用于中标结果公告等结构化页面）
    table_fields = extract_table_fields(html)

    result = {}
    # 优先用表格字段
    TABLE_KEY_MAP = {
        '建设单位名称': 'client',
        '建设单位': 'client',
        '招标人': 'client',
        '工程地点': 'site',
        '建设地点': 'site',
        '中标工期（天）': 'duration',
        '工期（天）': 'duration',
        '中标价格': 'budget',        # 中标结果公告中的价格
        '合同估算价': 'budget',
        '招标控制价': 'budget',
        '中标单位名称': 'winner',
        '中标单位': 'winner',
        '项目经理姓名：': 'pm',
        '项目经理': 'pm',
        '标段编号': 'section_id',
        '项目编号': 'project_id',
        '招标方式': 'bid_method',    # 复用字段
    }
    for table_key, field_key in TABLE_KEY_MAP.items():
        if table_key in table_fields:
            result[field_key] = table_fields[table_key]

    # 文本字段提取（适用于招标公告正文）
    text = clean_text(html)
    result.setdefault('client', '')
    result.setdefault('site', '')
    result.setdefault('budget', '')
    for kw, field in [('招标人：', 'client'), ('建设单位：', 'client'),
                      ('建设地点：', 'site'), ('工程地点：', 'site'),
                      ('合同估算价', 'budget'), ('招标控制价', 'budget')]:
        val = find_field(text, [kw], 150)
        if val and not result.get(field):
            result[field] = val

    if '资格后审' in text:
        result['qual_check'] = '资格后审'
    elif '资格预审' in text:
        result['qual_check'] = '资格预审'

    if '评标办法' in text:
        bm = find_field(text, ['综合评估法', '合理低价法', '经评审的最低投标价法', '评定分离'], 200)
        if bm:
            result['bid_method'] = bm

    result.setdefault('bid_method', '')
    result.setdefault('qual_check', '')
    result.setdefault('winner', table_fields.get('中标单位名称', ''))
    result.setdefault('pm', table_fields.get('项目经理姓名：', ''))
    result.setdefault('duration', table_fields.get('中标工期（天）', ''))

    return result

# ── PDF 字段提取（pymupdf）─────────────────────────────────────────────────
def extract_pdf_fields(pdf_path):
    """
    从招标文件 PDF 中提取关键字段。
    依赖 pymupdf（需提前 pip install pymupdf --break-system-packages）
    """
    try:
        import pymupdf
    except ImportError:
        return {}
    try:
        doc = pymupdf.open(pdf_path)
    except Exception:
        return {}

    full_text = ''
    for page in doc:
        full_text += page.get_text() + '\n'
    doc.close()

    text = re.sub(r'\s+', ' ', full_text).strip()
    # raw_text 保留原始换行；text_sp 用空格替换换行，用于跨行匹配
    raw_text = full_text
    text_sp = re.sub(r'\s+', ' ', raw_text)

    def _between(text, start_pat, end_pat, length=200):
        """在 start_pat 和 end_pat 之间提取内容（用于 raw_text 精确章节定位）"""
        m = re.search(start_pat, text)
        if not m:
            return ''
        start = m.end()
        # 找 end_pat（下一个章节编号）
        end_match = re.search(end_pat, text[start:])
        end = start + end_match.start() if end_match else start + length
        val = text[start:end].strip()
        # 清理：去掉换行和多余空白
        val = re.sub(r'\s+', ' ', val).strip()
        val = val.rstrip('。').rstrip('．').strip()
        return val[:length]

    def _find_after(text, keyword, length=200):
        """在文本中找关键词后的值（用于含空格的 text）"""
        idx = text.find(keyword)
        if idx == -1:
            return ''
        segment = text[idx:idx+length+len(keyword)]
        # 在第一个换行或连续空白处截断
        segment = re.sub(r'\n+', ' ', segment)
        val = segment[len(keyword):].strip()
        # 去掉前导冒号和空格
        val = re.sub(r'^[：:\s]+', '', val)
        return val[:length].rstrip('。').rstrip('．').strip()

    # ── 1. 项目名称 ──────────────────────────────────────────────────────
    project_name = _between(raw_text, r'2\.1\s*标段名称[：:\s*]', r'\n\s*\d+\.\d+', 80)

    # ── 2. 标段编号 ─────────────────────────────────────────────────────
    section_id = ''
    # 格式1: 标段编号：B3204051839000142001001
    m = re.search(r'标段编号[）)：:\s]*([A-Z0-9]{10,30})', raw_text)
    if m:
        section_id = m.group(1).strip()
    if not section_id:
        # 格式2: B3204811839000008039001（标段编号）
        m = re.search(r'([A-Z0-9]{10,30})[（\(]?\s*（?标段编号', raw_text)
        if m:
            section_id = m.group(1).strip()

    # ── 3. 招标人 ───────────────────────────────────────────────────────
    # 招标公告正文格式：招标人为 XXX
    client = _between(raw_text, r'招标人为', r'\n', 60)
    if not client:
        client = _between(raw_text, r'2\.1[^\n]*?招标人为', r'\n\s*\d+\.\d+', 60)
    if not client:
        client = _between(raw_text, r'招标人\s+名\s+称[：:\s*]', r'\n', 60)

    # ── 4. 招标代理机构 ─────────────────────────────────────────────────
    agent = _between(raw_text, r'招标代理机构[：:]\s*', r'\n', 60)

    # ── 5. 建设地点 ─────────────────────────────────────────────────────
    site = _between(raw_text, r'2\.2\s*建设地点[：:\s*]', r'\n\s*\d+\.\d+', 60)

    # ── 6. 建设内容 ─────────────────────────────────────────────────────
    content = _between(raw_text, r'2\.3\s*建设内容[：:\s*]', r'\n\s*\d+\.\d+', 120)

    # ── 7. 质量要求 ────────────────────────────────────────────────────
    quality = _between(raw_text, r'(?:2\.4|2\.5)\s*质量要求[：:\s*]', r'\n\s*\d+\.\d+', 30)

    # ── 8. 工程规模 ────────────────────────────────────────────────────
    scale = _between(raw_text, r'2\.5\s*工程规模[：:\s*]', r'\n\s*\d+\.\d+', 80)

    # ── 9. 工程合同估算价（万元） ─────────────────────────────────────────
    # 格式：工程合同估算价（万元）：717 或 工程合同估算价：717万元
    # 数字在冒号后面，可能有空格；也可能直接是"XXX万元"格式
    budget = ''
    idx_budget = text_sp.find('工程合同估算价')
    if idx_budget >= 0:
        chunk = text_sp[idx_budget:idx_budget + 60]
        m_num = re.search(r'[：:]\s*(\d+(?:\.\d+)?)', chunk)
        if m_num:
            budget = m_num.group(1).strip()
    if not budget:
        m2 = re.search(r'(\d+\.?\d*)\s*万元', text_sp)
        if m2:
            budget = m2.group(1).strip()

    # ── 10. 招标范围 ────────────────────────────────────────────────────
    # 精确章节提取（2.7 或 1.3 章节格式）
    scope = _between(raw_text, r'2\.7\s*(?:单位工程及\s*)?招标范围[：:\s*]', r'\n\s*\d+\.\d+', 100)
    if not scope:
        scope = _between(raw_text, r'1\.3\.1\s*本次招标范围[：:\s*]', r'\n', 100)
    if not scope:
        # 文档中通常写"见投标人须知前附表"，属于标准格式，直接引用
        scope = '见投标人须知前附表'

    # ── 11. 工期 ───────────────────────────────────────────────────────
    duration = ''
    m = re.search(r'要求工期[：:\s]*(\d+)\s*日历天', raw_text)
    if m:
        duration = m.group(1) + '日历天'

    # ── 12. 计划开工/竣工日期 ──────────────────────────────────────────
    start_date = ''
    m = re.search(r'计划开工日期[：:\s]*(\d{4}\s*年\s*\d{1,2}\s*月\s*\d{1,2}\s*日)', raw_text)
    if m:
        start_date = m.group(1).strip()
    end_date = ''
    m = re.search(r'计划竣工日期[：:\s]*(\d{4}\s*年\s*\d{1,2}\s*月\s*\d{1,2}\s*日)', raw_text)
    if m:
        end_date = m.group(1).strip()

    # ── 13. 投标人资质 ─────────────────────────────────────────────────
    qual_req = _between(raw_text, r'3\.1\s*投标人资质类别和等级[：:\s*]', r'\n\s*3\.2', 120)

    # ── 14. 项目负责人资质 ─────────────────────────────────────────────
    pm_req = _between(raw_text, r'3\.2\s*拟选派项目负责人专业及资质等级[：:\s*]', r'\n\s*3\.3', 100)

    # ── 15. 评标办法 ───────────────────────────────────────────────────
    # 15.1 主要评标办法（来自评标办法前附表）
    bid_method = ''
    if '☑合理低价法' in raw_text:
        bid_method = '合理低价法'
    elif '☑评定分离' in raw_text:
        bid_method = '综合评估法-评定分离'
    elif '☑综合评估法' in raw_text:
        bid_method = '综合评估法'
    elif '☑经评审的最低投标价法' in raw_text:
        bid_method = '经评审的最低投标价法'
    elif '☑经评审的最低投标价' in raw_text:
        bid_method = '经评审的最低投标价法'

    # 15.2 评标入围方法（第三章 评标办法 → 评标入围 → 2.1 评标入围标准）
    # 在"评标入围"章节中定位 ☑/□方法X：<name> 块，并提取"不少于 N 家"
    # 注意：部分文件用 ☑（选中）标记每种方法；部分文件用 □ + 前面"☑招标人直接确定以下方法X"的组合
    shortlist_method = ''
    shortlist_count = ''
    method_idx = text_sp.find('\u2611方法')   # ☑ ballot box with check
    if method_idx < 0:
        method_idx = text_sp.find('\u25a1方法')  # □ white square (部分文件用 □ 标记方法)
    if method_idx >= 0:
        # 向前找最近的"评标入围"正文区域
        section_start = text_sp.rfind('评标入围', 0, method_idx)
        if section_start < 0:
            section_start = text_sp.rfind('评标入围方法', 0, method_idx)
        if section_start >= 0:
            shortlist_section = text_sp[section_start:section_start+3000]
            # 匹配选中方法：☑方法X：name 或 □方法X：name
            m_short = re.search(r'[\u2611\u25a1]方法([一二三四五六])[：:]\s*(\S{2,10}(?:法|入围))', shortlist_section)
            if m_short:
                shortlist_method = m_short.group(2)
                # 精确定位本方法段落范围：从方法名开始，到下一个 ☑/□方法 或章节边界为止
                method_pos = m_short.start()
                method_start = max(0, method_pos - 50)
                # 找下一个方法标记（☑方法 / □方法），避免误捕获下一个方法的"不少于 X 家"
                next_method_m = re.search(r'[\u2611\u25a1]方法[一二三四五六]', shortlist_section[method_pos+20:])
                if next_method_m:
                    method_end = method_pos + 20 + next_method_m.start()
                else:
                    method_end = method_pos + 500  # fallback
                method_block = shortlist_section[method_start:method_end]
                m_count = re.search(r'不少于\s*(\d+)\s*家', method_block)
                if m_count:
                    shortlist_count = m_count.group(1)
    # 合并：主要方法 + 评标入围方法（数量）
    if shortlist_method and shortlist_count:
        bid_method = f'{bid_method}，{shortlist_method}（{shortlist_count}家）'
    elif shortlist_method:
        bid_method = f'{bid_method}，{shortlist_method}'

    # ── 15c. 综合评估法评分组成 ────────────────────────────────────────
    # 仅综合评估法项目，从 2.3.4 节提取各评审因素得分
    # 格式：施工组织设计12分+答辩2分+报价81分+信用分5分
    eval_detail = ''
    if '☑综合评估法' in raw_text or '☑综合评估法' in text_sp:
        parts_detail = []
        # 找 2.3.4 节
        idx_234 = text_sp.find('2.3.4 (1)')
        if idx_234 < 0:
            idx_234 = text_sp.find('2.3.4(1)')
        if idx_234 >= 0:
            sec_234 = text_sp[idx_234:idx_234 + 8000]
            # 2.3.4(1) 施工组织设计 + 答辩
            tidx = sec_234.find('评分因素 页数要求 分值')
            if tidx >= 0:
                tbl = sec_234[tidx:tidx + 1200]
                rows = re.split(r'☑', tbl)
                construct = 0
                for row in rows:
                    row = row.strip()
                    if not row:
                        continue
                    # 找每行末尾的 X 分
                    m_score = re.search(r'(\d+)\s*分', row)
                    if not m_score:
                        continue
                    score = int(m_score.group(1))
                    # 跳过：答辩（单独提取）、□其他（分值待定）、票决法等非施工组织设计项
                    # 只在该分数项"之前"的范围内查找 skip 关键字
                    # 避免 PDF 表格跨节时，后续小节内容被错误拼接到前一行
                    first_score_pos = row.find(m_score.group(0))
                    score_region = row[:first_score_pos + len(m_score.group(0))]
                    if any(k in score_region for k in ['答辩', '□其他', '定标委员会', '票决']):
                        continue
                    # 跳过页码（如"1 分 51"中的51）或超大数字
                    if score > 20:
                        continue
                    construct += score
                # 项目负责人答辩：从 sec_234 宽口径搜索"总分X 分"
                # tbl 末尾约 2077，答辩"总分2 分"在约 1551，需扩大范围
                def_chunk = sec_234[1170:1800]  # 覆盖答辩段落（含"总分"）
                m_def = re.search(r'总分\s*(\d+)\s*分', def_chunk)
                if m_def:
                    parts_detail.append(f'答辩{int(m_def.group(1))}分')
                if construct > 0:
                    parts_detail.insert(0, f'施工组织设计{construct}分')

            # 2.3.4(3) 报价评审
            idx3 = sec_234.find('2.3.4 (3)')
            if idx3 < 0:
                idx3 = sec_234.find('2.3.4(3)')
            if idx3 >= 0:
                chunk3 = sec_234[idx3:idx3 + 400]
                m3 = re.search(r'投标报价评审[^\d]*(\d+)\s*分', chunk3)
                if not m3:
                    m3 = re.search(r'(\d+)\s*分', chunk3[:100])
                if m3 and int(m3.group(1)) > 0:
                    parts_detail.append(f'报价{int(m3.group(1))}分')

            # 2.3.4(5) 信用分
            idx5 = sec_234.find('2.3.4 (5)')
            if idx5 < 0:
                idx5 = sec_234.find('2.3.4(5)')
            if idx5 >= 0:
                chunk5 = sec_234[idx5:idx5 + 400]
                # 优先找"×N/100" multiplier
                m_mult = re.search(r'×\s*(\d+)\s*/\s*100', chunk5)
                if m_mult:
                    parts_detail.append(f'信用分{int(m_mult.group(1))}分')
                else:
                    m5 = re.search(r'(\d+)\s*分', chunk5[:100])
                    if m5 and int(m5.group(1)) > 0:
                        parts_detail.append(f'信用分{int(m5.group(1))}分')

        eval_detail = '，'.join(parts_detail) if parts_detail else ''
        # 综合评估法：拼接评分组成到 bid_method
        if eval_detail:
            bid_method = f'{bid_method}，{eval_detail}'

    # ── 15b. 评标基准价计算方法 ────────────────────────────────────────
    # 从 2.3.2 评标基准价计算 方法 一节提取
    # 格式：ABC合成法，K（95%-98%），下浮率Δ（6%-12%）
    基准价计算 = ''
    sec_232 = text_sp.find('2.3.2 评标基准价计算')
    if sec_232 >= 0:
        section_232 = text_sp[sec_232:sec_232 + 4000]

        # 1) 选中的评标基准价方法编号（从 ☑...方法X 选择语句提取）
        cn_map = {
            '一':'方法一','二':'方法二','三':'方法三',
            '四':'方法四','五':'方法五','六':'方法六',
        }
        m_sel = re.search(r'☑.*?方法([一二三四五六]+)', section_232[:200])
        method_num = m_sel.group(1) if m_sel else ''
        method_key = cn_map.get(method_num, '')

        # 2) 方法名称（在 2.3.2 正文区域搜索该方法的正式名称，如 ABC合成法）
        method_name = ''
        if method_key:
            first_p = section_232.find(method_key)
            desc_p = section_232.find(method_key, first_p + 1) if first_p >= 0 else -1
            if desc_p < 0 or desc_p < 100:
                desc_p = first_p
            if desc_p >= 0:
                chunk = section_232[desc_p:desc_p + 80]
                chunk_ns = chunk.replace(' ', '')
                m_name = re.search(r'方法[一二三四五六]+[：:]*(\S*法)', chunk_ns)
                if m_name:
                    method_name = m_name.group(1)

        # 3) K 值范围（支持连续范围 95%～98% 和离散值 95%、96%、97%、98%）
        # 只在"取值范围"段落内提取，避免受后面 20% 等干扰
        k_area_m = re.search(r'K 值的取值范围[^\n；;]{0,60}', section_232[:2000])
        if k_area_m:
            k_vals = re.findall(r'(\d+(?:\.\d+)?)%', k_area_m.group())
            if k_vals:
                k_range = f"{k_vals[0]}%-{k_vals[-1]}%"
        if not k_range:
            k_range = ''

        # 4) 下浮率Δ分类 + 对应取值范围
        delta_range = ''
        tidx = section_232.find('本次招标项目下浮率Δ')
        if tidx >= 0:
            delta_block = section_232[tidx:tidx + 1000]
            # 下浮率Δ分类为
            m_class = re.search(r'下浮率Δ分类为\s*(\S+?)\s+(?=分类|$)', delta_block)
            delta_class_raw = m_class.group(1) if m_class else ''
            delta_class = delta_class_raw.rstrip('。./、')

            if delta_class and delta_class != '/':
                cat_key_map = {
                    '房屋建筑工程':'房屋建筑','市政工程':'市政',
                    '绿化工程':'绿化','机电安装工程':'机电安装',
                    '装饰装修工程':'装饰装修',
                }
                cat_key = cat_key_map.get(delta_class, delta_class.rstrip('工程'))
                # 下浮率Δ row（全部类别值连续排列）
                m_row = re.search(r'率Δ\s+(房屋建筑|装饰装修|机电安装|市政|绿化)', delta_block)
                if m_row:
                    row_start = m_row.start()
                    delta_row = delta_block[row_start:row_start + 600]
                    cat_pos = delta_row.find(cat_key)
                    if cat_pos >= 0:
                        # 找下一个类别名位置，确定本类别值的边界
                        # 找下一个类别名的位置
                        cat_positions = {c: delta_row.find(c) for c in ['房屋建筑','装饰装修','机电安装','市政','绿化']}
                        sorted_cats = [(c, p) for c, p in cat_positions.items() if p >= 0]
                        sorted_cats.sort(key=lambda x: x[1])
                        next_pos = len(delta_row)
                        for c, p in sorted_cats:
                            if p > cat_pos:
                                next_pos = p
                                break
                        vals = re.findall(r'(\d+)%', delta_row[cat_pos:next_pos])
                        if vals:
                            delta_range = f"{vals[0]}%-{vals[-1]}%"

        # 拼接最终字段值
        parts = []
        if method_name:
            parts.append(method_name)
        if k_range:
            parts.append(f"K（{k_range}）")
        parts.append(f"下浮率Δ（{delta_range if delta_range else '/'}）")
        基准价计算 = '，'.join(parts)

    # ── 16. 资格审查 ───────────────────────────────────────────────────
    qual_check = ''
    # "采用资格后审" 直接描述资格审查方式，最可靠
    if re.search(r'采用\s*资格后审', raw_text):
        qual_check = '资格后审'
    elif re.search(r'采用\s*资格预审', raw_text):
        qual_check = '资格预审'
    # 备用：PDF checkbox（U+2611 = ☑）
    elif re.search(r'\u2611\s*资格后审', raw_text):
        qual_check = '资格后审'
    elif re.search(r'\u2611\s*资格预审', raw_text):
        qual_check = '资格预审'

    # ── 17. 投标保证金 ────────────────────────────────────────────────
    deposit = ''
    # 匹配 "人民币5 万元" 或 "人民币 50 万元" 等格式
    m = re.search(r'人民币\s*(\d+\s*[^\s。]{0,10}万元?)', text_sp)
    if m:
        deposit = m.group(1).strip()[:40]

    # ── 18. 招标控制价/最高限价 ─────────────────────────────────────────
    bid_limit = ''
    m = re.search(r'最高投标限价\s+金额[：:\s]*([^\n]{2,60})', raw_text)
    if m:
        val = m.group(1).strip()
        if '详见' in val:
            bid_limit = '详见公布的招标控制价'
        else:
            bid_limit = val[:60]
    if not bid_limit:
        m = re.search(r'招标控制价[：:\s]*([^\n]{2,60})', raw_text)
        if m:
            val = m.group(1).strip()
            if '详见' in val:
                bid_limit = '详见公布的招标控制价'
            else:
                bid_limit = val[:60]

    # ── 19. 合同价格形式 ───────────────────────────────────────────────
    price_type = ''
    if '☑单价合同' in raw_text:
        price_type = '单价合同'
    elif '☑总价合同' in raw_text:
        price_type = '总价合同'

    # ── 20. 投标有效期 ─────────────────────────────────────────────────
    valid_period = ''
    m = re.search(r'投标有效期\s+(\d+)\s*天', raw_text)
    if m:
        valid_period = m.group(1) + '天'

    # ── 21. 履约担保 ──────────────────────────────────────────────────
    perform_guarantee = ''
    # 搜索"合同总价/合同金额...的X%"模式（标准格式）
    # 同时兼容 ASCII % 和全角 ％
    m = re.search(
        r'(?:合同总价|合同金额)[^。，\n]{0,80}?的\s*\d+\s*[％%]',
        raw_text
    )
    if m:
        val = m.group(0).strip()
        # 去除尾部残余页码和章节号
        val = re.sub(r'\s*[.。…]+\s*\d+\s*$', '', val)
        val = re.sub(r'\s+\d+\.\d+\s*$', ' ', val)
        perform_guarantee = val[:80]
    # 备用：直接搜索"合同总价的5"类模式（朱林邻里案例）
    if not perform_guarantee:
        m2 = re.search(r'(?:合同总价|合同金额)[^。，\n]{0,100}?的\s*\d+\s*[％%]', raw_text)
        if m2:
            perform_guarantee = m2.group(0).strip()[:80]

    # ── 22. 付款方式（工程款支付方式）──────────────────────────────────
    payment = _between(raw_text, r'工程款支付方式[：:\s]*', r'\n', 80)

    # ── 23. 投标截止时间 ───────────────────────────────────────────────
    bid_deadline = _between(raw_text, r'投标截止时间为[：:\s]*', r'\n', 60)

    # ── 24. 招标人联系方式 ──────────────────────────────────────────────
    # 策略：优先在"10.联系方式" / "招标主体" 区域搜索（该区域同时含招标人+代理联系信息）
    contact = ''
    agent_contact = ''
    # 定位"联系方式"章节（通常在文档后半部）
    contact_section_start = text_sp.rfind('10.联系方式')
    if contact_section_start < 0:
        contact_section_start = text_sp.rfind('招标主体')
    if contact_section_start < 0:
        contact_section_start = text_sp.rfind('联系方式')
    if contact_section_start >= 0:
        # 取该章节后的 ~800 字符（足够包含招标人+代理联系信息）
        contact_section = text_sp[contact_section_start:contact_section_start+800]

        # 招标人电话：在"招标人"附近找电话
        m = re.search(r'招标人[^人在]{0,30}(?:电话[:：]\s*)?(\d{3,4}[-\s]?\d{7,8})', contact_section)
        if m:
            contact = '招标人 ' + m.group(1)
        # 代理电话
        m2 = re.search(r'招标代理机构[^机构]{0,30}(?:电话[:：]\s*)?(\d{3,4}[-\s]?\d{7,8})', contact_section)
        if m2:
            agent_contact = '代理 ' + m2.group(1)
        # 如果代理没单独找到，尝试在 contact_section 找所有电话
        if not agent_contact:
            phones = re.findall(r'(?<![-\d])(\d{3,4}[-\s]?\d{7,8})(?![-\d])', contact_section)
            if len(phones) >= 2:
                contact = f'招标人 {phones[0]}'
                agent_contact = f'代理 {phones[1]}'
            elif phones:
                contact = f'招标人 {phones[0]}'

    # 备用：在全文范围用宽松模式（适用于联系人在同一行的情况）
    if not contact:
        m = re.search(r'招标人[^\n]*?(?:电话[:：]\s*)?(\d{3,4}[-\s]?\d{7,8})', raw_text)
        if m:
            contact = '招标人 ' + m.group(1)

    # ── 25. 代理联系方式 ────────────────────────────────────────────────
    if not agent_contact:
        m = re.search(r'招标代理机构[^\n]*?(?:电话[:：]\s*)?(\d{3,4}[-\s]?\d{7,8})', raw_text)
        if m:
            agent_contact = '代理 ' + m.group(1)

    # ── 26. 是否接受联合体 ─────────────────────────────────────────────
    joint_bid = ''
    if '☑不接受' in raw_text and '联合体' in raw_text:
        joint_bid = '不接受联合体'
    elif '☑接受' in raw_text and '联合体' in raw_text:
        joint_bid = '接受联合体'

    return {
        'project_name': project_name,
        'section_id': section_id,
        'client': client,
        'agent': agent,
        'site': site,
        'content': content,
        'quality': quality,
        'scale': scale,
        'budget': budget,
        'bid_limit': bid_limit,
        'scope': scope,
        'duration': duration,
        'start_date': start_date,
        'end_date': end_date,
        'qual_req': qual_req,
        'pm_req': pm_req,
        'bid_method': bid_method,
        'qual_check': qual_check,
        'deposit': deposit,
        'perform_guarantee': perform_guarantee,
        'price_type': price_type,
        'valid_period': valid_period,
        'payment': payment,
        'bid_deadline': bid_deadline,
        'contact': contact,
        'agent_contact': agent_contact,
        'joint_bid': joint_bid,
        '基准价计算': 基准价计算,
        'eval_detail': eval_detail,
    }


# ── PDF 附件提取+下载 ───────────────────────────────────────────────────────
def fetch_pdfs(html, out_dir):
    """
    从详情页 HTML 中提取 PDF 附件信息并下载。
    只下载"招标文件正文"（主要招标文件），不下载"招标公告"等次要附件。
    返回: [{'name': 文件名, 'size': 字节数, 'path': 本地路径}, ...]
    """
    results = []
    SKIP_KEYWORDS = ['招标公告', '定稿', '变更', '澄清', '补充']
    # 匹配: onclick="ztbfjyz('/path?attachGuid=...','1','0')" title="xxx.pdf"
    pattern = re.compile(
        r'onclick="ztbfjyz\(\'([^\']+)\',\s*\'1\',\s*\'0\'\)"[^>]*title="([^"]+\.pdf)"'
    )
    for m in pattern.finditer(html):
        pdf_path, fname = m.group(1), m.group(2)
        # 跳过招标公告等次要文件，只下载招标文件正文
        if any(kw in fname for kw in SKIP_KEYWORDS) and '招标文件正文' not in fname:
            print(f'      跳过(次要): {fname}')
            continue
        guid_m = re.search(r'attachGuid=([a-f0-9\-]+)', pdf_path)
        if not guid_m:
            continue
        guid = guid_m.group(1)
        pdf_url = BASE + f'/czggzyweb/WebbuilderMIS/attach/downloadZtbAttach.jspx?attachGuid={guid}&appUrlFlag=ztb007&siteGuid={SITE_GUID}'
        safe_fname = re.sub(r'[\\/:*?"<>|]', '_', fname)
        out_path = os.path.join(out_dir, safe_fname)
        try:
            size = download(pdf_url, out_path)
            results.append({'name': fname, 'size': size, 'path': out_path})
            print(f'      PDF: {fname} ({size:,} bytes)')
        except Exception as e:
            print(f'      PDF Error [{fname}]: {e}')
    return results

# ── 项目分类判断 ────────────────────────────────────────────────────────────
def get_sub_cat(cat):
    """根据 cat 末3位判断二级类型"""
    code = cat[-3:] if len(cat) >= 3 else cat
    return TYPE_NAMES.get(code, code)

TYPE_NAMES = {
    '001': '施工', '002': '服务', '003': '货物', '004': '货物',
    '005': '中标候选人公示', '006': '中标结果公告',
    '008': '中标合同', '009': '合同变更',
}

def cat_matches_filter(cat, filter_cat):
    """
    判断 cat 是否符合 filter_cat 过滤条件。
    cat12 格式: [主类别6位][子类别3位][类型末3位]
    示例: 001001 | 001 | 001 = 建设工程-招标公告-施工
          001001 | 006 | 001 = 建设工程-中标结果公告-施工

    filter_cat=001001001001 → 前9位=001001001（招标公告主类）AND 末3位=001（施工）
    filter_cat=001001001002 → 前9位=001001001 AND 末3位=002（服务）
    filter_cat=001001005   → 前6位=001001 且 cat[6:9]=005（中标候选人主类）
    """
    if len(cat) < 3 or len(filter_cat) < 3:
        return True

    # 9位 cat: 只有末3位
    if len(cat) == 9:
        return cat[-3:] == filter_cat[-3:]

    # 12位 cat: 比较前9位(主类+子类) 和末3位(类型)
    cat_main = cat[:9]
    cat_type = cat[-3:]
    f_main = filter_cat[:min(len(filter_cat), 9)]
    f_type = filter_cat[-3:]

    # 对于精确到施工/服务/货物的请求（前9位已知）
    if len(f_main) == 9:
        main_match = cat_main == f_main
    else:
        # 前6位匹配（主类别）
        main_match = cat[:6] == filter_cat[:6]

    type_match = cat_type == f_type
    return main_match and type_match

# ═══════════════════════════════════════════════════════════════════════════════
# 自动复盘：每次运行后核查字段提取质量 → 更新 changelog
# ═══════════════════════════════════════════════════════════════════════════════
def review_and_log_changelog(results, run_label=""):
    """
    对 results 中的每条记录核查关键字段（PDF原文 vs 提取值），
    生成简要质量报告并追加到 changelog.md。
    """
    import pymupdf, re as _re, io as _io
    from datetime import datetime

    VALIDATION_FIELDS = [
        'budget', 'bid_method', '基准价计算', 'qual_check',
        'deposit', 'duration', 'qual_req', 'pm_req',
        'price_type', 'valid_period', 'perform_guarantee',
    ]

    report_lines = []
    total = 0
    ok_count = 0

    for idx, it in enumerate(results):
        pdfs = it.get('pdfs', [])
        if not pdfs:
            continue
        main_pdf = None
        for pf in pdfs:
            if '招标文件正文' in pf.get('name', ''):
                main_pdf = pf['path']
                break
        if not main_pdf:
            main_pdf = pdfs[0]['path']

        title = it.get('title', it.get('project_name', f'条目{idx+1}'))[:40]

        try:
            doc = pymupdf.open(main_pdf)
            full = '\n'.join(p.get_text() for p in doc)
            doc.close()
        except Exception:
            continue

        text_sp = _re.sub(r'\s+', ' ', full)
        extracted = extract_pdf_fields(main_pdf)  # 复用已有提取函数

        total += 1
        item_ok = 0
        item_fields = []

        for field in VALIDATION_FIELDS:
            val = extracted.get(field, '')
            if val and val.strip():
                item_ok += 1
                status = '✅'
            elif field in ('perform_guarantee',):  # 部分字段可为空
                status = 'N/A'
            else:
                status = '❌'

            item_fields.append(f"  {status} {field}: {val!r}")

        # ── 综合评估法总分=100 检查 ──────────────────────────────────────────
        bid_method_val = extracted.get('bid_method', '')
        if '综合评估法' in bid_method_val:
            # 提取所有 X分 模式
            scores = _re.findall(r'(\d+)\s*分', bid_method_val)
            if scores:
                total_score = sum(int(s) for s in scores)
                if total_score == 100:
                    item_fields.append(f"  ✅ 综合评估法总分: {total_score}分（符合要求）")
                    item_ok += 1   # 这项通过才计入
                else:
                    item_fields.append(f"  ⚠️  综合评估法总分: {total_score}分（应为100分，不匹配！）")
                    # 不 +1，分数不达标

        ok_count += item_ok

        report_lines.append(f"**{title}**")
        report_lines += item_fields

    # ── 写入 changelog ─────────────────────────────────────────────────────
    score = f"{ok_count}/{total * len(VALIDATION_FIELDS)}" if total > 0 else "N/A"
    date_str = datetime.now().strftime('%Y-%m-%d')

    entry = f"""
### {date_str} {run_label} {'✅' if '❌' not in chr(10).join(report_lines) else '⚠️'} 评分：{score}
"""
    if report_lines:
        entry += "**抓取条目**：\n" + '\n'.join(report_lines) + "\n"

    changelog_path = os.path.join(os.path.dirname(__file__), 'changelog.md')

    # 读取现有内容，插入到"历史运行记录"之后（第一个 ### 标题处）
    try:
        with open(changelog_path, 'r', encoding='utf-8') as f:
            content = f.read()
    except Exception:
        content = ""

    # 在 "## 历史运行记录" 之后插入
    marker = "## 历史运行记录"
    if marker in content:
        parts = content.split(marker, 1)
        entry = entry.strip()
        content = parts[0] + marker + "\n\n---\n\n" + entry + "\n\n" + parts[1].lstrip("\n")
    else:
        content = content + "\n\n" + entry

    with open(changelog_path, 'w', encoding='utf-8') as f:
        f.write(content)

    print(f"\n[字段质量复盘] {'✅ 无问题' if '❌' not in entry else '⚠️ 有字段未提取'} | {score}")
    return ok_count, total


# ── 持续积累 Excel 读写（openpyxl）──────────────────────────────────────────
def read_persistent_excel():
    """
    读取本地持续积累 Excel，返回 (data_list, existing_section_ids_set)
    data_list: [{'section_id':..., 'date':..., 'uuid':..., ...}, ...]
    失败时返回 ([], set())
    """
    try:
        import openpyxl
    except ImportError:
        return [], set()

    path = PERSISTENT_EXCEL
    if not os.path.exists(path):
        return [], set()

    try:
        wb = openpyxl.load_workbook(path)  # 不用 data_only=True，避免 pandas 写的文件读回 None
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        # 建立列名→索引映射
        h_map = {h: i for i, h in enumerate(headers) if h}

        sid_idx = h_map.get('标段编号', -1)
        date_idx = h_map.get('发布日期', -1)
        uuid_idx = h_map.get('来源', -1)   # placeholder, we'll track uuid differently

        # uuid 在当前 Excel 中没有独立列，靠文件名目录关联
        # 读取 section_id 作为去重键
        existing_sids = set()
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 处理空单元格：pandas/openpyxl 写空字符串后读回 None，统一转为 ''
            row_clean = [cell if cell is not None else '' for cell in row]
            sid = row_clean[sid_idx] if sid_idx >= 0 else ''
            if sid:
                existing_sids.add(str(sid).strip())
            rows.append(dict(zip(headers, row_clean)))

        print(f'    [持久Excel] 读取 {len(rows)} 条历史记录')
        return rows, existing_sids
    except Exception as e:
        print(f'    [WARN] 读取持久Excel失败: {e}')
        return [], set()


def write_persistent_excel(all_items):
    """
    将 all_items（已去重+已过滤过期）写入固定文件名 PERSISTENT_EXCEL。
    """
    # 全局表头顺序
    headers = ['序号', '项目名称', '地区', '发布日期', '分类',
               '招标人', '招标代理', '建设地点', '建设内容',
               '控制价(万元)', '评标办法', '评标基准价计算方法', '资格审查',
               '投标保证金', '工期', '计划开工', '计划竣工',
               '资质要求', '项目负责人资质', '合同价格形式', '投标有效期',
               '履约担保', '联系方式', '标段编号', '来源']

    # 写入 openpyxl
    try:
        import openpyxl
    except ImportError:
        make_xlsx(all_items, PERSISTENT_EXCEL)
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '招标项目明细'

    # 写表头
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)

    # 写数据
    def _fv(it, k_en, k_cn=''):
        """读取字段，兼容英文 key（标准化后）和中文 key（历史 Excel）"""
        return it.get(k_en) or (it.get(k_cn) if k_cn else '') or ''

    for ri_offset, it in enumerate(all_items, 2):
        row_vals = [
            str(ri_offset - 1),
            _fv(it, 'project_name', '项目名称'),
            _fv(it, 'area', '地区'),
            _fv(it, 'date', '发布日期'),
            _fv(it, 'sub_cat', '分类'),
            _fv(it, 'client', '招标人'),
            _fv(it, 'agent', '招标代理'),
            _fv(it, 'site', '建设地点'),
            _fv(it, 'content', '建设内容')[:300],
            _fv(it, 'budget', '控制价(万元)'),
            _fv(it, 'bid_method', '评标办法'),
            _fv(it, '基准价计算', '评标基准价计算方法'),
            _fv(it, 'qual_check', '资格审查'),
            _fv(it, 'deposit', '投标保证金'),
            _fv(it, 'duration', '工期'),
            _fv(it, 'start_date', '计划开工'),
            _fv(it, 'end_date', '计划竣工'),
            _fv(it, 'qual_req', '资质要求'),
            _fv(it, 'pm_req', '项目负责人资质'),
            _fv(it, 'price_type', '合同价格形式'),
            _fv(it, 'valid_period', '投标有效期'),
            _fv(it, 'perform_guarantee', '履约担保'),
            _fv(it, 'contact', '联系方式') or _fv(it, 'agent_contact', ''),
            _fv(it, 'section_id', '标段编号'),
            '常州公共资源交易中心',
        ]
        for ci, v in enumerate(row_vals, 1):
            ws.cell(row=ri_offset, column=ci, value=str(v) if v != '' else '')

    os.makedirs(os.path.dirname(PERSISTENT_EXCEL), exist_ok=True)
    wb.save(PERSISTENT_EXCEL)
    print(f'  持续Excel写入: {PERSISTENT_EXCEL}  ({len(all_items)} 条)')


def cleanup_expired_pdf_folders(items_to_remove):
    """
    根据过期条目（section_id 匹配）删除本地 PDF 文件夹。
    PDF 文件夹命名格式: {date_str}_{uuid_prefix}
    items_to_remove: [{'section_id':..., 'date':..., 'title':...}, ...]
    """
    # 从 outputs 目录找出所有 yyyyMMdd_* 格式的文件夹
    if not os.path.isdir(OUT_DIR):
        return
    for fname in os.listdir(OUT_DIR):
        fpath = os.path.join(OUT_DIR, fname)
        if not os.path.isdir(fpath):
            continue
        # 匹配格式: {date}_{uuid_prefix} 或直接含 section_id
        for it in items_to_remove:
            sid = it.get('section_id', '')
            if sid and sid in fname:
                shutil.rmtree(fpath)
                print(f'    删除过期PDF目录: {fname}')
                break


def merge_and_cleanup(new_items):
    """
    读取本地持久 Excel，合并 new_items，过滤过期条目，删除过期PDF，写回。
    返回: 合并后的全部有效条目列表
    """
    print(f'\n[Step 3½] 持久Excel增量合并（保留近 {MAX_DAYS} 天）...')

    # 1. 读取历史
    existing_rows, existing_sids = read_persistent_excel()

    # 2. 构建历史字典（key=section_id），用于去重和快速查找
    #    也保留 uuid 匹配（PDF 文件夹名含 uuid prefix）
    hist_map = {}   # section_id → row_dict
    for row in existing_rows:
        sid = row.get('标段编号', '')
        if sid:
            hist_map[str(sid).strip()] = row

    # 3. 合并新增（去重：新数据优先，因为可能更新了字段）
    merged = dict(hist_map)
    for it in new_items:
        sid = it.get('section_id', '') or it.get('标段编号', '')
        if sid and sid in merged:
            # 已存在：合并字段（新值覆盖空值，但保留旧值不为空的内容）
            old = merged[sid]
            for k, v in it.items():
                # 只覆盖空值；new_item 的 None/空 不覆盖 old 的有效值
                if v and v not in ('', 'NULL', 'None') and (old.get(k) in (None, '', 'NULL', 'None') or k not in old):
                    old[k] = v
            # 补齐 area/date/title 等列表页字段
            for fk in ['area', 'date', 'title', 'uuid', 'sub_cat', 'project_name']:
                if fk in it and it[fk] and it[fk] not in ('', 'NULL', 'None') and (old.get(fk) in (None, '', 'NULL', 'None') or fk not in old):
                    old[fk] = it[fk]
            # ── 统一 section_id / 标段编号 key ──────────────────────────────
            if 'section_id' in it and it['section_id']:
                old['section_id'] = it['section_id']
                if '标段编号' not in old or not old.get('标段编号'):
                    old['标段编号'] = it['section_id']
            elif '标段编号' in old and old['标段编号']:
                old['section_id'] = old['标段编号']
        else:
            # 新增
            merged[sid if sid else f"NEW_{it.get('uuid','')}"] = it
            if 'section_id' not in it and '标段编号' in it:
                it['section_id'] = it['标段编号']

    # ── 合并后整体标准化（确保所有 item 的 key 统一，方便后续读写 Excel）──
    for key, item in merged.items():
        # 标准化中文列名 → 英文 key（make_xlsx 用英文 key 读写）
        if '项目名称' in item and 'project_name' not in item:
            item['project_name'] = item['项目名称']
        if '标段编号' in item and 'section_id' not in item:
            item['section_id'] = item['标段编号']
        if '发布日期' in item and 'date' not in item:
            item['date'] = item['发布日期']
        if '地区' in item and 'area' not in item:
            item['area'] = item['地区']
        if '分类' in item and 'sub_cat' not in item:
            item['sub_cat'] = item['分类']
        # ── 补全其余字段的中文→英文映射（腾讯文档写入用英文 key）────────────
        for cn, en in [
            ('招标人',           'client'),
            ('招标代理',         'agent'),
            ('建设地点',         'site'),
            ('建设内容',         'content'),
            ('控制价(万元)',     'budget'),
            ('评标办法',         'bid_method'),
            ('评标基准价计算方法', '基准价计算'),
            ('资格审查',        'qual_check'),
            ('投标保证金',      'deposit'),
            ('工期',            'duration'),
            ('计划开工',        'start_date'),
            ('计划竣工',        'end_date'),
            ('资质要求',        'qual_req'),
            ('项目负责人资质',   'pm_req'),
            ('合同价格形式',    'price_type'),
            ('投标有效期',      'valid_period'),
            ('履约担保',        'perform_guarantee'),
            ('联系方式',        'contact'),
        ]:
            if cn in item and en not in item:
                item[en] = item[cn]
        # 统一 section_id / 标段编号
        sid = item.get('section_id', '') or item.get('标段编号', '')
        if sid:
            item['section_id'] = sid
            item['标段编号'] = sid

    # 4. 过滤过期（以当天为基准，>7天删除）
    #    注意：有 section_id 但 date 为 None 的历史遗留条目（网站端无日期）→ 保留（不因 date 丢失而误删）
    today = datetime.now().date()
    valid_items = []
    expired_items = []
    for sid, it in merged.items():
        d = parse_date(it.get('date', ''))
        if d is None:
            # 无日期，默认保留
            valid_items.append(it)
            continue
        if d >= today - timedelta(days=MAX_DAYS):
            valid_items.append(it)
        else:
            expired_items.append(it)

    expired_count = len(expired_items)
    print(f'    合并: {len(merged)} 条 → 有效: {len(valid_items)} 条 | 过期删除: {expired_count} 条')
    if expired_count > 0:
        for e in expired_items:
            print(f'      - {e.get("date","")} | {e.get("title",e.get("project_name",""))[:40]}')

    # 5. 删除过期 PDF 文件夹
    if expired_items:
        cleanup_expired_pdf_folders(expired_items)

    # 6. 按日期倒序
    valid_items.sort(key=lambda x: x.get('date', ''), reverse=True)

    # 7. 写回持久 Excel
    write_persistent_excel(valid_items)

    return valid_items
def make_xlsx(data, out_path):
    headers = ['序号', '项目名称', '地区', '发布日期', '分类',
               '招标人', '招标代理', '建设地点', '建设内容',
               '控制价(万元)', '评标办法', '评标基准价计算方法', '资格审查',
               '投标保证金', '工期', '计划开工', '计划竣工',
               '资质要求', '项目负责人资质', '合同价格形式', '投标有效期',
               '履约担保', '联系方式', '标段编号', '来源']
    col_widths = [5, 40, 10, 12, 10,
                  24, 24, 24, 38,
                  14, 16, 20, 10,
                  16, 12, 14, 14,
                  40, 30, 12, 10,
                  30, 30, 30, 22]

    def cr(c, r):
        """将列索引(0-based)转为Excel列字母（如0->A, 25->Z, 26->AA）"""
        s = ''
        c += 1  # 转为 1-based
        while c:
            c, rem = divmod(c - 1, 26)
            s = chr(65 + rem) + s
        return s + str(r)
    def esc(s): return str(s).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('"','&quot;').replace("'", '&apos;')

    strings, si_map = [], {}
    def si(s):
        s = str(s)
        if s not in si_map:
            si_map[s] = len(strings)
            strings.append(s)
        return si_map[s]

    for h in headers: si(h)

    def _fv(d, k_en, k_cn=''):
        return d.get(k_en) or (d.get(k_cn) if k_cn else '') or ''

    def row_fields(d, ri):
        pdfs = d.get('pdfs', [])
        pdf_list = '; '.join(p['name'] for p in pdfs) if pdfs else ''

        return [
            str(ri-1),
            _fv(d, 'project_name', '项目名称'),
            _fv(d, 'area', '地区'),
            _fv(d, 'date', '发布日期'),
            _fv(d, 'sub_cat', '分类'),
            _fv(d, 'client', '招标人'),
            _fv(d, 'agent', '招标代理'),
            _fv(d, 'site', '建设地点'),
            _fv(d, 'content', '建设内容')[:300],
            _fv(d, 'budget', '控制价(万元)'),
            _fv(d, 'bid_method', '评标办法'),
            _fv(d, '基准价计算', '评标基准价计算方法'),
            _fv(d, 'qual_check', '资格审查'),
            _fv(d, 'deposit', '投标保证金'),
            _fv(d, 'duration', '工期'),
            _fv(d, 'start_date', '计划开工'),
            _fv(d, 'end_date', '计划竣工'),
            _fv(d, 'qual_req', '资质要求'),
            _fv(d, 'pm_req', '项目负责人资质'),
            _fv(d, 'price_type', '合同价格形式'),
            _fv(d, 'valid_period', '投标有效期'),
            _fv(d, 'perform_guarantee', '履约担保'),
            _fv(d, 'contact', '联系方式') or _fv(d, 'agent_contact', ''),
            _fv(d, 'section_id', '标段编号'),
            '常州公共资源交易中心',
        ]

    rows_xml = ['<row r="1">' + ''.join(f'<c r="{cr(i,1)}" s="1" t="s"><v>{si(h)}</v></c>' for i,h in enumerate(headers)) + '</row>']
    for ri, d in enumerate(data, 2):
        fields = row_fields(d, ri)
        for v in fields: si(v)
        rows_xml.append('<row r="'+str(ri)+'">' + ''.join(f'<c r="{cr(i,ri)}" s="2" t="s"><v>{si(v)}</v></c>' for i,v in enumerate(fields)) + '</row>')

    cols = ''.join(f'<col min="{i+1}" max="{i+1}" width="{w}" customWidth="1"/>' for i,w in enumerate(col_widths))
    sheet   = f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetViews><sheetView workbookViewId="0" showGridLines="1"><selection activeCell="A1" sqref="A1"/></sheetView></sheetViews><cols>{cols}</cols><sheetData>{"".join(rows_xml)}</sheetData><pageMargins left="0.5" right="0.5" top="0.75" bottom="0.75"/></worksheet>'
    ss_xml  = f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="{len(strings)}" uniqueCount="{len(strings)}">{"".join(f"<si><t>{esc(s)}</t></si>" for s in strings)}</sst>'
    wb      = f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><sheets><sheet name="\u62db\u6807\u9879\u76ee\u660e\u7ec6" sheetId="1" r:id="rId1"/></sheets></workbook>'
    styles  = f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><fonts><font><sz val="11"/><name val="\u5fae\u8f6f\u96c5\u9ed1"/></font><font><sz val="11"/><b/><name val="\u5fae\u8f6f\u96c5\u9ed1"/></font></fonts><fills><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill><fill><patternFill patternType="solid"><fgColor rgb="FF028090"/></patternFill></fill></fills><borders><border><left/><right/><top/><bottom/><diagonal/></border></borders><cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs><cellXfs><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/><xf numFmtId="0" fontId="1" fillId="2" borderId="0" xfId="0"><alignment horizontal="center"/></xf><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"><alignment wrapText="1"/></xf></cellXfs></styleSheet>'
    wb_rels = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/><Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/><Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/></Relationships>'
    root_rels = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/></Relationships>'
    ct      = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/><Default Extension="xml" ContentType="application/xml"/><Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/><Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/><Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/><Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/></Types>'

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as z:
        z.writestr('[Content_Types].xml', ct)
        z.writestr('_rels/.rels', root_rels)
        z.writestr('xl/workbook.xml', wb)
        z.writestr('xl/_rels/workbook.xml.rels', wb_rels)
        z.writestr('xl/styles.xml', styles)
        z.writestr('xl/worksheets/sheet1.xml', sheet)
        z.writestr('xl/sharedStrings.xml', ss_xml)

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, 'wb') as f:
        f.write(buf.getvalue())
    print(f'  Excel: {out_path}  ({os.path.getsize(out_path)} bytes)')

# ── 分类配置 ────────────────────────────────────────────────────────────────
CATEGORIES = {
    '001001':       '建设工程-不限',
    '001001001':    '建设工程-招标公告/资审公告',
    '001001001001': '建设工程-招标公告-施工',   # 默认
    '001001001002': '建设工程-招标公告-服务',
    '001001001004': '建设工程-招标公告-货物',
    '001001005':    '建设工程-中标候选人/评标结果公示',
    '001001006':    '建设工程-中标结果公告',
    '001001008':    '建设工程-中标合同',
    '001001009':    '建设工程-合同履行及变更',
    '031':          '建设工程-交易变更公告',
    '001002':       '交通工程',
    '001003':       '水利工程',
    '001004':       '政府采购',
    '001005':       '土地矿产',
    '001006':       '国有产权',
    '001009':       '其他交易',
}

# ── 主函数 ─────────────────────────────────────────────────────────────────
def main():
    args = sys.argv[1:]
    # 解析可选参数
    sync = '--sync' in args
    notify = '--notify' in args
    dry_run = '--dry-run' in args
    if sync: args.remove('--sync')
    if notify: args.remove('--notify')
    if dry_run: args.remove('--dry-run')

    cat    = args[0] if len(args) > 0 else '001001001001'
    max_n  = int(args[3]) if len(args) > 3 else 20
    out_dir = OUT_DIR

    # ── 自动7天滑动窗口 ───────────────────────────────────────────────────
    today = datetime.now().strftime('%Y-%m-%d')
    week_ago = (datetime.now() - timedelta(days=MAX_DAYS)).strftime('%Y-%m-%d')
    start = week_ago    # 始终抓近7天
    end   = today

    cat_name = CATEGORIES.get(cat, cat)

    print('='*60)
    print('  常州市公共资源交易中心招标信息抓取工具 v2')
    print('  分类：' + cat_name)
    if sync:
        print('  模式：同步腾讯文档（仅保留近 {} 天）'.format(MAX_DAYS))
    print('='*60)
    print(f'  日期: {start} ~ {end} | 最大: {max_n}\n')

    print('[Step 1] 抓取常州列表页（近{}天）...'.format(MAX_DAYS))
    html = get(BASE + '/jyzx/001001/tradeInfonew.html?category=' + cat)
    print(f'  HTML: {len(html)} bytes')

    items = parse_list(html)
    print(f'  解析: {len(items)} 条\n')

    if not items:
        print('ERROR: 0 items. Website structure may have changed.')
        return

    # 按日期窗口过滤（近7天）
    filt = [x for x in items if x['date'] >= start and x['date'] <= end]
    # 按 cat 末3位过滤（列表页 category 参数无效，按 cat 实际类型筛选）
    filt = [x for x in filt if cat_matches_filter(x['cat'], cat)]
    filt.sort(key=lambda x: x['date'], reverse=True)

    print(f'[Result] {len(filt)} items（在 {start} ~ {end} 窗口内）')
    for i, it in enumerate(filt[:8]):
        print(f'  [{i+1}] {it["date"]} | {it["area"]} | {it["title"][:48]}')
    print()

    to_proc = filt[:max_n]
    print(f'[Step 2] 抓取详情页 + PDF ({len(to_proc)} items)...')
    results = []
    for i, it in enumerate(to_proc):
        print(f'  [{i+1}/{len(to_proc)}] {it["title"][:40]}...')

        # 获取详情页 URL（AJAX）
        detail_url = get_detail_url(it['uuid'], it['cat'])
        if not detail_url:
            print(f'      [WARN] 无法获取详情页 URL，跳过')
            it['sub_cat'] = get_sub_cat(it['cat'])
            it.update({k: '' for k in ['client','site','budget','bid_method','qual_check','qual_req','corp_perf','open_date','tech_bid','enter_method','payment','duration','phone','deposit','pdfs']})
            results.append(it)
            continue

        # 抓详情页
        try:
            detail_html = get(detail_url)
        except Exception as e:
            print(f'      [WARN] 详情页请求失败: {e}')
            it['sub_cat'] = get_sub_cat(it['cat'])
            it.update({k: '' for k in ['client','site','budget','bid_method','qual_check','qual_req','corp_perf','open_date','tech_bid','enter_method','payment','duration','phone','deposit','pdfs']})
            results.append(it)
            continue

        # 提取字段
        fields = extract_detail_fields(detail_html)
        it['sub_cat'] = get_sub_cat(it['cat'])
        for k, v in fields.items():
            if v:
                it[k] = v

        # 下载 PDF 并提取字段
        date_str = it['date'].replace('-', '')
        pdf_out_dir = os.path.join(out_dir, f'{date_str}_{it["uuid"][:8]}')
        os.makedirs(pdf_out_dir, exist_ok=True)
        it['pdfs'] = fetch_pdfs(detail_html, pdf_out_dir)

        # 从 PDF 提取关键字段（覆盖 HTML 提取的字段）
        if it['pdfs']:
            first_pdf = it['pdfs'][0]['path']
            pdf_fields = extract_pdf_fields(first_pdf)
            for k, v in pdf_fields.items():
                if v:
                    it[k] = v
            if pdf_fields:
                print(f'      PDF字段提取: {list(pdf_fields.keys())[:6]}...')

        results.append(it)
        time.sleep(1.0)

    # ── 自动复盘：核查字段提取质量并更新 changelog ──────────────────────────
    review_and_log_changelog(results, run_label="第三次运行")

    # ── Step 3: 持久Excel增量合并（去重+过期清理） ─────────────────────────
    # 合并本地历史记录，删除>7天过期项及对应PDF，生成全量有效列表
    merged_all = merge_and_cleanup(results)

    # 打印汇总（基于本次抓取的 items）
    total_pdfs = sum(len(r.get('pdfs',[])) for r in results)
    print(f'\n[DONE] 本次抓取 {len(results)} 条，PDF {total_pdfs} 个')
    print(f'       持续积累 Excel 现有 {len(merged_all)} 条（已合并去重+过期清理）')
    for r in results:
        pdfs = r.get('pdfs', [])
        print(f'  {r["date"]} | {r["sub_cat"]} | {len(pdfs)} PDFs | {r["title"][:40]}')

    # ── Step 4: 腾讯文档同步（基于合并后的全量数据）────────────────────────
    if sync:
        new_count, del_count = qq_sync('mcporter', merged_all, dry_run=dry_run)
        if not dry_run:
            print(f'\n  同步完成：新增 {new_count} 项，删除过期 {del_count} 项')
            print(f'  腾讯文档：{TENCENT_URL}')
        else:
            print(f'\n  [DRY RUN] 同步完成：新增 {new_count} 项，删除过期 {del_count} 项')

    # ── Step 5: 邮件通知（资质匹配推送）───────────────────────────────────
    if notify:
        print('\n[Step 5] 招标邮件通知（资质匹配）...')
        import importlib.util, sys as _sys
        # 动态加载，避免顶层依赖
        spec = importlib.util.spec_from_file_location(
            "qualification_matcher",
            str(Path(__file__).parent.parent / "companies" / "qualification_matcher.py")
        )
        qm = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(qm)

        spec2 = importlib.util.spec_from_file_location(
            "bid_email_sender",
            str(Path(__file__).parent / "bid_email_sender.py")
        )
        bes = importlib.util.module_from_spec(spec2)
        spec2.loader.exec_module(bes)

        companies = qm.load_companies()
        if not companies:
            print('  [WARN] 公司数据库为空，请先在 companies/companies.json 中添加企业信息')
        else:
            print(f'  公司数据库: {len(companies)} 家企业')
            from collections import defaultdict
            # 按公司分组匹配结果
            email_groups = defaultdict(list)  # email → [matched_items]

            for it in results:
                qual_req = it.get('qual_req', '') or ''
                if not qual_req:
                    continue
                matched = qm.match_companies_to_bid(companies, qual_req)
                for company in matched:
                    email = company.get('email', '')
                    if email:
                        email_groups[email].append(it)

            if not email_groups:
                print('  本次抓取项目中无匹配企业的公告，跳过通知')
            else:
                print(f'  将向 {len(email_groups)} 个企业发送邮件')
                for email_addr, items in email_groups.items():
                    print(f'    → {email_addr} ({len(items)} 条)')
                if not dry_run:
                    for email_addr, items in email_groups.items():
                        bes.send_bid_notifications(items, recipients=[email_addr], dry_run=False)
                else:
                    print('  [DRY RUN] 跳过实际发送')

if __name__ == '__main__':
    main()
