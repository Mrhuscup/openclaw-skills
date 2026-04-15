#!/usr/bin/env python3
"""生成常州招标 Excel - 纯净版"""
import urllib.request, ssl, re, io, zipfile, os
from datetime import datetime

BASE = 'http://ggzy.xzsp.changzhou.gov.cn'

def get_html(url):
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    with urllib.request.urlopen(req, timeout=15, context=ctx) as r:
        return r.read().decode('utf-8', 'ignore')

# ── 抓取并解析列表页 ──
html = get_html(BASE + '/jyzx/001001/tradeInfonew.html?category=001001')
print('HTML:', len(html), 'bytes')

REL = ['市政', '道路', '公路', '桥梁', '排水', '照明', '绿化', '交通',
       '施工', '总承包', '新建', '改建']
score = lambda t: sum(1 for k in REL if k in t)

items = []
seen = set()
for m in re.finditer(r'<tr[^>]*>([\s\S]*?)</tr>', html):
    row = m.group(1)
    segs = row.split('</td>')
    if len(segs) < 4:
        continue
    om = re.search(r"tzjydetail\('([^']+)',\s*'([^']+)'\s*,", segs[1])
    if not om or om.group(2) in segs[0]:
        continue
    tm = re.search(r'title="([^"]{5,120})"', segs[1])
    if not tm:
        continue
    uuid = om.group(2)
    if uuid in seen:
        continue
    seen.add(uuid)
    area = re.sub(r'<[^>]+>', '', segs[2]).replace('\xa0', ' ').strip()
    date = re.sub(r'<[^>]+>', '', segs[3]).replace('\xa0', ' ').strip()
    items.append({'cat': om.group(1), 'uuid': uuid,
                  'title': tm.group(1)[:120], 'area': area, 'date': date})

print('解析:', len(items), '条')

# 过滤 + 排序
filt = [x for x in items if '2026-03-01' <= x['date'] <= '2026-04-02']
filt.sort(key=lambda x: (score(x['title']), x['date']), reverse=True)
top = filt[:3]

print('过滤后 Top3:')
for i, it in enumerate(top):
    print(f'  {i+1}. [{it["date"]}] {it["area"]} | {it["title"][:55]}')

# ── 生成 Excel（openpyxl + 正确格式）───
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '招标项目明细'

    hdr_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='028090')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell_align = Alignment(wrap_text=True, vertical='top')

    headers = ['序号', '项目名称', '建设单位', '控制价（万元）',
               '入围方法', '评标办法', '开标日期', '开标时间',
               '付款方式', '资质要求', '企业业绩要求',
               '项目经理业绩要求', '是否编写技术标', '保证金（万元）']
    col_widths = [6, 52, 28, 14, 20, 24, 13, 10, 42, 42, 42, 42, 16, 16]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(1, ci, h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 32

    for ri, row in enumerate(top, 2):
        fields = [str(ri-1), row['title'], '', '', '', '',
                  row['date'], '', '', '', '', '', '', '']
        for ci, val in enumerate(fields, 1):
            c = ws.cell(ri, ci, val)
            c.alignment = cell_align
            c.border = border
        ws.row_dimensions[ri].height = 44

    out = '/workspace/skills/changzhou-bid-search/outputs/常州招标_2026-04-01.xlsx'
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    sz = os.path.getsize(out)
    print(f'\nOK: {out}  ({sz} bytes)')

except ImportError:
    # 降级：手动拼 xlsx（不用 sharedStrings，用 inlineStr）
    def cr(col, row):
        return chr(65 + col) + str(row)

    def esc(s):
        s = str(s)
        return (s.replace('&', '&amp;').replace('<', '&lt;')
                   .replace('>', '&gt;').replace('"', '&quot;'))

    headers = ['序号', '项目名称', '建设单位', '控制价（万元）',
               '入围方法', '评标办法', '开标日期', '开标时间',
               '付款方式', '资质要求', '企业业绩要求',
               '项目经理业绩要求', '是否编写技术标', '保证金（万元）']
    col_widths = [6, 52, 28, 14, 20, 24, 13, 10, 42, 42, 42, 42, 16, 16]

    rows_xml = [
        '<row r="1">' +
        ''.join(f'<c r="{cr(i,1)}" s="1" t="inlineStr"><is><t>{esc(h)}</t></is></c>'
                for i, h in enumerate(headers)) +
        '</row>'
    ]
    for ri, row in enumerate(top, 2):
        fields = [str(ri-1), row['title'], '', '', '', '',
                  row['date'], '', '', '', '', '', '', '']
        rows_xml.append(
            '<row r="' + str(ri) + '">' +
            ''.join(f'<c r="{cr(i,ri)}" s="2" t="inlineStr"><is><t>{esc(v)}</t></is></c>'
                    for i, v in enumerate(fields)) +
            '</row>'
        )

    cols_xml = ''.join(
        f'<col min="{i+1}" max="{i+1}" width="{w}" customWidth="1"/>'
        for i, w in enumerate(col_widths)
    )

    sheet = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<sheetViews>'
        '<sheetView workbookViewId="0" showGridLines="1">'
        '<selection activeCell="A1" sqref="A1"/>'
        '</sheetView>'
        '</sheetViews>'
        '<cols>' + cols_xml + '</cols>'
        '<sheetData>' + ''.join(rows_xml) + '</sheetData>'
        '<pageMargins left="0.5" right="0.5" top="0.75" bottom="0.75"/>'
        '</worksheet>'
    )

    styles = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<fonts>'
        '<font><sz val="11"/><name val="微软雅黑"/></font>'
        '<font><sz val="11"/><b/><name val="微软雅黑"/></font>'
        '</fonts>'
        '<fills>'
        '<fill><patternFill patternType="none"/></fill>'
        '<fill><patternFill patternType="gray125"/></fill>'
        '<fill><patternFill patternType="solid"><fgColor rgb="FF028090"/></fgColor></patternFill></fill>'
        '</fills>'
        '<borders><border><left/><right/><top/><bottom/><diagonal/></border></borders>'
        '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
        '<cellXfs>'
        '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>'
        '<xf numFmtId="0" fontId="1" fillId="2" borderId="0" xfId="0"><alignment horizontal="center"/></xf>'
        '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"><alignment wrapText="1"/></xf>'
        '</cellXfs>'
        '</styleSheet>'
    )

    wb_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheets><sheet name="招标项目明细" sheetId="1" r:id="rId1"/></sheets>'
        '</workbook>'
    )

    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
        '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
        '</Relationships>'
    )

    ct = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
        '</Types>'
    )

    root = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        '</Relationships>'
    )

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as z:
        z.writestr('[Content_Types].xml', ct)
        z.writestr('_rels/.rels', root)
        z.writestr('xl/workbook.xml', wb_xml)
        z.writestr('xl/_rels/workbook.xml.rels', rels)
        z.writestr('xl/styles.xml', styles)
        z.writestr('xl/worksheets/sheet1.xml', sheet)

    out = '/workspace/skills/changzhou-bid-search/outputs/常州招标_2026-04-01.xlsx'
    os.makedirs(os.path.dirname(out), exist_ok=True)
    with open(out, 'wb') as f:
        f.write(buf.getvalue())
    sz = os.path.getsize(out)
    print(f'\nOK (fallback): {out}  ({sz} bytes)')
