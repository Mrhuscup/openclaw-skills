#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
供应商资质自动更新工作流 v5.3 - 持续日志版

功能：
1. 从腾讯云文档导出完整供应商库Excel（本地已有则跳过）
2. 筛选30天内到期和已过期的供应商
3. 住建部官网查询最新安全许可证有效期
4. 更新本地Excel文件
5. 在同一个腾讯云文档中持续追加日志

用法：
    python supplier_update_workflow.py [--days 30] [--limit N]
"""

import sys
import os
import json
import time
import subprocess
import shutil
from datetime import datetime, timedelta
from pathlib import Path

# 添加scripts路径
sys.path.insert(0, str(Path(__file__).parent))

from mohurd_scraper import MohurdScraper
from email_sender import send_supplier_update_report

# 配置
TENCENT_DOCS_FILE_ID = "XvdgsQTsPGdZ"
OUTPUT_DIR = Path("/root/.openclaw/workspace-supplier")
LOG_DIR = Path("/root/.openclaw/workspace-supplier/logs")
LOCAL_EXCEL = "/root/.openclaw/workspace/supplier_full.xlsx"
LOG_DOC_CONFIG = OUTPUT_DIR / "log_doc_config.json"

# 腾讯云文档日志配置
TENCENT_LOG_DOC_TITLE = "供应商更新日志"

# mcporter 环境变量
def get_mcporter_env():
    """获取 mcporter 所需的环境变量"""
    env = os.environ.copy()
    env['PATH'] = '/root/.nvm/versions/node/v22.22.2/bin:' + env.get('PATH', '')
    return env


class TencentDocsLogger:
    """腾讯云文档日志记录器 - 持续更新版"""
    
    def __init__(self, title=TENCENT_LOG_DOC_TITLE):
        self.title = title
        self.doc_id = None
        self.sheet_id = None
        self.current_row = 1
        self.config_file = LOG_DOC_CONFIG
        
    def _run_mcporter(self, func_name, args):
        """执行 mcporter 命令"""
        args_str = json.dumps(args) if args else '{}'
        
        # 设置环境变量，确保能找到 mcporter
        env = os.environ.copy()
        env['PATH'] = '/root/.nvm/versions/node/v22.22.2/bin:' + env.get('PATH', '')
        
        result = subprocess.run(
            ['mcporter', 'call', 'tencent-docs', func_name, '--args', args_str],
            capture_output=True, text=True, env=env
        )
        if result.returncode != 0:
            print(f"[ERROR] mcporter failed: {result.stderr}")
            return None
        try:
            return json.loads(result.stdout)
        except:
            return None
    
    def load_existing_doc(self):
        """加载已存在的日志文档配置"""
        if self.config_file.exists():
            with open(self.config_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
                self.doc_id = config.get('doc_id')
                self.sheet_id = config.get('sheet_id')
                self.current_row = config.get('current_row', 1)
                return True
        return False
    
    def save_doc_config(self):
        """保存日志文档配置"""
        config = {
            'doc_id': self.doc_id,
            'sheet_id': self.sheet_id,
            'current_row': self.current_row,
            'title': self.title,
            'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        with open(self.config_file, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
    
    def init_doc(self):
        """初始化日志文档（创建新或复用已有）"""
        # 尝试加载已存在的文档
        if self.load_existing_doc() and self.doc_id:
            # 验证文档是否还存在
            result = self._run_mcporter('manage.query_file_info', {
                'file_id': self.doc_id
            })
            if result and not result.get('error'):
                print(f"[INFO] 复用已有日志文档: {self.doc_id}")
                # 获取当前行数
                self._refresh_current_row()
                return True
            else:
                print(f"[INFO] 已有文档已失效，将创建新文档")
                self.doc_id = None
                self.sheet_id = None
        
        # 创建新文档
        return self.create_new_doc()
    
    def create_new_doc(self):
        """创建新的日志文档"""
        result = self._run_mcporter('manage.create_file', {
            'file_type': 'sheet',
            'title': self.title
        })
        
        if not result or result.get('error'):
            print(f"[ERROR] 创建腾讯云文档失败: {result}")
            return False
        
        self.doc_id = result.get('file_id')
        doc_url = result.get('url', '')
        print(f"[INFO] 创建腾讯云文档成功: {doc_url}")
        
        # 获取 sheet ID
        sheet_result = self._run_mcporter('sheet.operation_sheet', {
            'file_id': self.doc_id,
            'js_script': 'const sheets = SpreadsheetApp.getActiveSpreadsheet().getSheets(); console.log(JSON.stringify(sheets.map(s => ({name: s.getName(), id: s.getSheetId()}))));'
        })
        
        if sheet_result and 'log' in sheet_result:
            import re
            match = re.search(r'"id":"([^"]+)"', sheet_result['log'])
            if match:
                self.sheet_id = match.group(1)
        
        if not self.sheet_id:
            print(f"[ERROR] 获取sheet_id失败")
            return False
        
        # 写入表头
        self._write_header()
        
        # 保存配置
        self.save_doc_config()
        
        return True
    
    def _refresh_current_row(self):
        """刷新当前行数"""
        result = self._run_mcporter('sheet.operation_sheet', {
            'file_id': self.doc_id,
            'sheet_id': self.sheet_id,
            'js_script': f'const sheet = SpreadsheetApp.getActiveSpreadsheet().getSheetById("{self.sheet_id}"); console.log(sheet.getLastRow());'
        })
        
        if result and 'log' in result:
            try:
                self.current_row = int(result['log'].strip()) + 1
            except:
                self.current_row = 2
    
    def _write_header(self):
        """写入表头 - 【优化P2】合并为单次调用"""
        if not self.doc_id or not self.sheet_id:
            return
        
        # 【优化P2】合并7次setValue为一次脚本执行
        js_script = f'''
        const spreadsheet = SpreadsheetApp.getActiveSpreadsheet();
        const sheet = spreadsheet.getSheetById("{self.sheet_id}");
        const headers = ["序号","执行时间","公司名称","原日期","新日期","状态","批次"];
        headers.forEach((h, i) => sheet.getRange(1, i+1).setValue(h));
        sheet.setColumnWidth(1, 60);
        sheet.setColumnWidth(2, 160);
        sheet.setColumnWidth(3, 300);
        sheet.setColumnWidth(4, 120);
        sheet.setColumnWidth(5, 120);
        sheet.setColumnWidth(6, 80);
        sheet.setColumnWidth(7, 200);
        '''
        
        self._run_mcporter('sheet.operation_sheet', {
            'file_id': self.doc_id,
            'sheet_id': self.sheet_id,
            'js_script': js_script
        })
        
        self.current_row = 2
    
    def append_update(self, index, company, old_date, new_date, batch="", status="已更新"):
        """追加更新记录"""
        if not self.doc_id or not self.sheet_id:
            return
        
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        company_escaped = company.replace("'", "\\'")
        
        js_script = f'''
        const spreadsheet = SpreadsheetApp.getActiveSpreadsheet();
        const sheet = spreadsheet.getSheetById("{self.sheet_id}");
        
        const row = {self.current_row};
        sheet.getRange(row, 1).setValue({index});
        sheet.getRange(row, 2).setValue("{timestamp}");
        sheet.getRange(row, 3).setValue("{company_escaped}");
        sheet.getRange(row, 4).setValue("{old_date}");
        sheet.getRange(row, 5).setValue("{new_date}");
        sheet.getRange(row, 6).setValue("{status}");
        sheet.getRange(row, 7).setValue("{batch}");
        '''
        
        self._run_mcporter('sheet.operation_sheet', {
            'file_id': self.doc_id,
            'sheet_id': self.sheet_id,
            'js_script': js_script
        })
        
        self.current_row += 1
    
    def append_no_data(self, index, company, current_date, batch="", note="无查询结果"):
        """追加无结果记录"""
        if not self.doc_id or not self.sheet_id:
            return
        
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        company_escaped = company.replace("'", "\\'")
        
        js_script = f'''
        const spreadsheet = SpreadsheetApp.getActiveSpreadsheet();
        const sheet = spreadsheet.getSheetById("{self.sheet_id}");
        
        const row = {self.current_row};
        sheet.getRange(row, 1).setValue({index});
        sheet.getRange(row, 2).setValue("{timestamp}");
        sheet.getRange(row, 3).setValue("{company_escaped}");
        sheet.getRange(row, 4).setValue("{current_date}");
        sheet.getRange(row, 5).setValue("-");
        sheet.getRange(row, 6).setValue("待处理");
        sheet.getRange(row, 7).setValue("{batch}");
        '''
        
        self._run_mcporter('sheet.operation_sheet', {
            'file_id': self.doc_id,
            'sheet_id': self.sheet_id,
            'js_script': js_script
        })
        
        self.current_row += 1
    
    def append_batch_header(self, batch_name, total_count):
        """追加批次标题行"""
        if not self.doc_id or not self.sheet_id:
            return
        
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        js_script = f'''
        const spreadsheet = SpreadsheetApp.getActiveSpreadsheet();
        const sheet = spreadsheet.getSheetById("{self.sheet_id}");
        
        const row = {self.current_row};
        sheet.getRange(row, 1).setValue("【{batch_name}】");
        sheet.getRange(row, 2).setValue("{timestamp}");
        sheet.getRange(row, 3).setValue("共计 {total_count} 条");
        '''
        
        self._run_mcporter('sheet.operation_sheet', {
            'file_id': self.doc_id,
            'sheet_id': self.sheet_id,
            'js_script': js_script
        })
        
        self.current_row += 1

    def append_all_records(self, updates, no_data, global_start_index=1, batch_name=""):
        """【优化P1】批量写入所有记录 - 一次调用完成所有写入
        大幅减少 API 调用次数，从 N+M+1 次降至 1 次
        """
        if not self.doc_id or not self.sheet_id:
            return
        
        # 【防覆盖修复】写入前先查询实际最后一行，确保不覆盖已有数据
        self._refresh_current_row()
        
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # 构造所有单元格写入脚本
        row = self.current_row
        js_parts = []
        
        # 1. 批次标题行
        js_parts.append(f'sheet.getRange({row}, 1).setValue("【{batch_name}】");')
        js_parts.append(f'sheet.getRange({row}, 2).setValue("{timestamp}");')
        js_parts.append(f'sheet.getRange({row}, 3).setValue("共计 {len(updates)+len(no_data)} 条");')
        row += 1
        
        # 2. 更新记录
        for u in updates:
            idx = global_start_index + (row - self.current_row)
            company_escaped = u['company'].replace('"', '\\"').replace("'", "\\'")
            js_parts.append(f'sheet.getRange({row}, 1).setValue({idx});')
            js_parts.append(f'sheet.getRange({row}, 2).setValue("{timestamp}");')
            js_parts.append(f'sheet.getRange({row}, 3).setValue("{company_escaped}");')
            js_parts.append(f'sheet.getRange({row}, 4).setValue("{u["old_date"]}");')
            js_parts.append(f'sheet.getRange({row}, 5).setValue("{u["new_date"]}");')
            js_parts.append(f'sheet.getRange({row}, 6).setValue("已更新");')
            js_parts.append(f'sheet.getRange({row}, 7).setValue("{batch_name}");')
            row += 1
        
        # 3. 无结果记录
        for item in no_data:
            idx = global_start_index + (row - self.current_row)
            company_escaped = item['company'].replace('"', '\\"').replace("'", "\\'")
            js_parts.append(f'sheet.getRange({row}, 1).setValue({idx});')
            js_parts.append(f'sheet.getRange({row}, 2).setValue("{timestamp}");')
            js_parts.append(f'sheet.getRange({row}, 3).setValue("{company_escaped}");')
            js_parts.append(f'sheet.getRange({row}, 4).setValue("{item["current_date"]}");')
            js_parts.append(f'sheet.getRange({row}, 5).setValue("-");')
            js_parts.append(f'sheet.getRange({row}, 6).setValue("待处理");')
            js_parts.append(f'sheet.getRange({row}, 7).setValue("{batch_name}");')
            row += 1
        
        js_script = f'''
        const spreadsheet = SpreadsheetApp.getActiveSpreadsheet();
        const sheet = spreadsheet.getSheetById("{self.sheet_id}");
        {'; '.join(js_parts)}
        '''
        
        # 【优化P1核心】一次调用写入全部！
        result = self._run_mcporter('sheet.operation_sheet', {
            'file_id': self.doc_id,
            'sheet_id': self.sheet_id,
            'js_script': js_script
        })
        
        if result and not result.get('error'):
            print(f"[优化] 批量写入{len(updates)+len(no_data)+1}行，仅用1次API调用")
        else:
            print(f"[ERROR] 批量写入失败: {result}")
        
        # 本地更新行号，不再查询文档
        self.current_row = row
    
    def finalize(self, total_updates, total_no_data):
        """完成写入，保存配置"""
        self.save_doc_config()
    
    def get_doc_url(self):
        """获取文档URL"""
        if self.doc_id:
            return f"https://docs.qq.com/sheet/{self.doc_id}"
        return None


class SupplierUpdateWorkflow:
    """供应商更新工作流"""
    
    def __init__(self, days=30, limit=None):
        self.days = days
        self.limit = limit
        self.today = datetime.now()
        self.cutoff = self.today + timedelta(days=days)
        self.batch_name = self.today.strftime('%Y%m%d_%H%M%S')
        
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        LOG_DIR.mkdir(parents=True, exist_ok=True)
        
        self.scraper = None
        self.tencent_logger = None
        
    def log(self, message, level="INFO"):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"[{timestamp}] [{level}] {message}")
        
    def export_excel_from_tencent_docs(self):
        self.log("步骤1: 从腾讯云文档下载最新供应商库...")
        
        # 【注意】每次执行都重新下载最新版本，确保使用最新数据
        # 不再使用本地缓存，因为腾讯云文档可能已被用户修改
        self.log("强制刷新：跳过本地缓存，从腾讯云文档获取最新文件...")
        
        result = subprocess.run([
            'mcporter', 'call', 'tencent-docs', 'manage.export_file',
            '--args', json.dumps({"file_id": TENCENT_DOCS_FILE_ID})
        ], capture_output=True, text=True, env=get_mcporter_env())
        
        if result.returncode != 0:
            self.log(f"导出失败: {result.stderr}", "ERROR")
            return None
        
        try:
            task_data = json.loads(result.stdout)
            task_id = task_data.get('task_id')
            
            if not task_id:
                self.log("获取task_id失败", "ERROR")
                return None
            
            self.log(f"等待导出完成 (task_id: {task_id})...")
            
            # 【优化P0】指数退避轮询：1+2+3+5+8+10+15+20+30 = 最多9次，比固定60次节省85%
            delays = [1, 2, 3, 5, 8, 10, 15, 20, 30]
            total_waited = 0
            
            for i, d in enumerate(delays):
                time.sleep(d)
                total_waited += d
                
                progress_result = subprocess.run([
                    'mcporter', 'call', 'tencent-docs', 'manage.export_progress',
                    '--args', json.dumps({"task_id": task_id})
                ], capture_output=True, text=True, env=get_mcporter_env())
                
                if progress_result.returncode == 0:
                    progress_data = json.loads(progress_result.stdout)
                    progress = progress_data.get('progress', 0)
                    file_url = progress_data.get('file_url', '')
                    
                    if progress == 100 and file_url:
                        # 保存到 LOCAL_EXCEL 路径，供后续步骤使用
                        subprocess.run([
                            'curl', '-s', '-o', LOCAL_EXCEL, file_url
                        ], check=True)
                        self.log(f"导出成功 (耗时{total_waited}秒): {LOCAL_EXCEL}")
                        return LOCAL_EXCEL
                
                self.log(f"等待中... ({total_waited}秒)")
            
            self.log(f"导出超时（已等待{total_waited}秒）", "ERROR")
                    
        except Exception as e:
            self.log(f"导出异常: {e}", "ERROR")
            
        return None
    
    def load_and_filter_suppliers(self, excel_file):
        import openpyxl
        
        self.log(f"步骤2: 加载并筛选供应商...")
        
        wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
        ws = wb.active
        
        company_col = 3
        safety_col = 20
        
        need_update = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            company = row[company_col] if len(row) > company_col else None
            safety_date = row[safety_col] if len(row) > safety_col else None
            
            if not company or not isinstance(company, str) or len(company) < 4:
                continue
            if '公司' not in company:
                continue
            
            if not safety_date:
                continue
                
            if isinstance(safety_date, datetime):
                date_obj = safety_date
            elif isinstance(safety_date, str):
                try:
                    date_obj = datetime.strptime(safety_date.split()[0], '%Y-%m-%d')
                except:
                    continue
            else:
                continue
            
            days = (date_obj - self.today).days
            if days <= self.days:
                need_update.append({
                    'row': row_idx,
                    'company': company,
                    'current_date': date_obj.strftime('%Y-%m-%d'),
                    'days': days,
                    'status': '已过期' if days < 0 else f'还有{days}天'
                })
        
        wb.close()
        
        self.log(f"找到 {len(need_update)} 条需要查询的供应商")
        
        seen = set()
        unique_list = []
        for item in need_update:
            if item['company'] not in seen:
                seen.add(item['company'])
                unique_list.append(item)
        
        self.log(f"去重后 {len(unique_list)} 条")
        
        if self.limit:
            unique_list = unique_list[:self.limit]
            self.log(f"限制查询前 {len(unique_list)} 条 (测试模式)")
        
        return unique_list
    
    def query_mohurd(self, company):
        try:
            if self.scraper is None:
                self.scraper = MohurdScraper(headless=True)
                self.scraper.start()
            
            return self.scraper.query(company, timeout=60)
        except Exception as e:
            self.log(f"查询异常: {company} - {e}", "WARNING")
            return None
    
    def update_local_excel(self, updates):
        import openpyxl
        
        self.log(f"步骤4: 更新本地Excel...")
        
        if not updates:
            self.log("没有需要更新的记录")
            return
        
        input_file = LOCAL_EXCEL
        output_file = LOCAL_EXCEL.replace('.xlsx', '_updated.xlsx')
        
        shutil.copy(input_file, output_file)
        
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        safety_col = 21
        
        for update in updates:
            row = update['row']
            new_date = update['new_date']
            
            try:
                old_date = ws.cell(row=row, column=safety_col).value
                ws.cell(row=row, column=safety_col).value = datetime.strptime(new_date, '%Y-%m-%d')
                self.log(f"  行{row}: {update['company'][:30]}... {old_date} -> {new_date}")
            except Exception as e:
                self.log(f"  更新失败: 行{row} - {e}", "ERROR")
        
        wb.save(output_file)
        wb.close()
        
        self.log(f"更新完成: {output_file}")
        
        return output_file
    
    def save_local_log(self, updates, no_data):
        log_file = LOG_DIR / f"update_log_{self.batch_name}.json"
        
        log_data = {
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'batch': self.batch_name,
            'total_updates': len(updates),
            'total_no_data': len(no_data),
            'updates': updates,
            'no_data': no_data
        }
        
        with open(log_file, 'w', encoding='utf-8') as f:
            json.dump(log_data, f, ensure_ascii=False, indent=2)
        
        self.log(f"本地日志已保存: {log_file}")
        
        return log_file
    
    def init_tencent_logger(self):
        """初始化腾讯云文档日志"""
        self.log("步骤5: 初始化腾讯云文档日志...")
        
        self.tencent_logger = TencentDocsLogger()
        
        if not self.tencent_logger.init_doc():
            self.log("初始化腾讯云文档失败", "ERROR")
            return False
        
        doc_url = self.tencent_logger.get_doc_url()
        self.log(f"日志文档: {doc_url}")
        
        return True
    
    def save_tencent_log(self, updates, no_data):
        """【优化P1】保存到腾讯云文档 - 使用批量写入"""
        if not self.tencent_logger:
            return None
        
        self.log(f"写入腾讯云文档 (批量模式)...")
        
        # 计算全局序号起始值
        global_start_index = self.tencent_logger.current_row
        
        # 【优化P1】一次调用写入所有记录
        self.tencent_logger.append_all_records(
            updates=updates,
            no_data=no_data,
            global_start_index=global_start_index,
            batch_name=self.batch_name
        )
        
        # 完成
        self.tencent_logger.finalize(len(updates), len(no_data))
        
        doc_url = self.tencent_logger.get_doc_url()
        self.log(f"腾讯云文档已更新: {doc_url}")
        
        return doc_url
    
    def send_email_notification(self, updates, no_data, tencent_url=None):
        """发送邮件通知"""
        self.log("步骤6: 发送邮件通知...")
        
        try:
            # 生成报告文本
            report_lines = []
            report_lines.append(f"供应商资质更新报告")
            report_lines.append(f"=" * 50)
            report_lines.append(f"执行时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            report_lines.append(f"批次号: {self.batch_name}")
            report_lines.append(f"")
            report_lines.append(f"更新数量: {len(updates)} 条")
            report_lines.append(f"待处理数量: {len(no_data)} 条")
            report_lines.append(f"")
            
            if updates:
                report_lines.append("【已更新】")
                for u in updates:
                    report_lines.append(f"  ✓ {u['company']}")
                    report_lines.append(f"    {u['old_date']} → {u['new_date']}")
                report_lines.append("")
            
            if no_data:
                report_lines.append("【待处理】")
                for item in no_data:
                    report_lines.append(f"  ✗ {item['company']} (原日期: {item['current_date']})")
                report_lines.append("")
            
            if tencent_url:
                report_lines.append(f"腾讯云文档日志: {tencent_url}")
            
            report_text = "\n".join(report_lines)
            
            # 发送邮件
            success = send_supplier_update_report(
                updates=updates,
                report_text=report_text,
                report_extra=f"批次: {self.batch_name}"
            )
            
            if success:
                self.log("邮件发送成功")
            else:
                self.log("邮件发送失败", "WARNING")
            
            return success
            
        except Exception as e:
            self.log(f"邮件发送异常: {e}", "ERROR")
            return False
    
    def run(self):
        self.log("=" * 60)
        self.log("供应商资质自动更新工作流 v5.4")
        self.log("=" * 60)
        
        try:
            # 1. 获取Excel
            excel_file = self.export_excel_from_tencent_docs()
            if not excel_file:
                self.log("获取Excel失败，退出", "ERROR")
                return
            
            # 2. 筛选供应商
            suppliers = self.load_and_filter_suppliers(excel_file)
            if not suppliers:
                self.log("没有需要更新的供应商", "INFO")
                return
            
            # 3. 查询住建部
            self.log("步骤3: 查询住建部官网...")
            
            updates = []
            no_data = []
            
            for i, supplier in enumerate(suppliers, 1):
                self.log(f"[{i}/{len(suppliers)}] 查询: {supplier['company']}")
                
                result = self.query_mohurd(supplier['company'])
                
                if result:
                    self.log(f"  结果: {result}")
                    updates.append({
                        'company': supplier['company'],
                        'row': supplier['row'],
                        'old_date': supplier['current_date'],
                        'new_date': result,
                        'source': '住建部官网'
                    })
                else:
                    self.log(f"  结果: 无数据", "WARNING")
                    no_data.append(supplier)
                
                time.sleep(2)
            
            # 4. 更新Excel
            if updates:
                self.update_local_excel(updates)
            
            # 5. 双渠道日志
            local_log = self.save_local_log(updates, no_data)
            tencent_url = None
            if self.init_tencent_logger():
                tencent_url = self.save_tencent_log(updates, no_data)
            
            # 6. 邮件通知
            email_success = False
            if updates:
                email_success = self.send_email_notification(updates, no_data, tencent_url)
            
            self.log("=" * 60)
            self.log("工作流完成!")
            self.log(f"更新数量: {len(updates)}")
            self.log(f"待处理数量: {len(no_data)}")
            self.log(f"本地日志: {local_log}")
            if tencent_url:
                self.log(f"腾讯云文档: {tencent_url}")
            self.log(f"邮件通知: {'成功' if email_success else '失败/跳过'}")
            self.log("=" * 60)
            
            return updates, no_data
            
        finally:
            if self.scraper:
                self.scraper.close()


def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='供应商资质自动更新工作流')
    parser.add_argument('--days', type=int, default=30, help='提前多少天预警 (默认30)')
    parser.add_argument('--limit', type=int, default=None, help='限制查询数量 (用于测试)')
    
    args = parser.parse_args()
    
    workflow = SupplierUpdateWorkflow(days=args.days, limit=args.limit)
    workflow.run()


if __name__ == "__main__":
    main()
