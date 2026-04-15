#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
供应商信息自动更新工作流 v5.0 - 整合真实爬虫版

更新说明：
- v5.0: 整合住建部官网真实爬虫，替换模拟数据逻辑
- 标招网因需要会员账号，暂不接入
- 使用 Playwright 实现浏览器自动化查询

依赖安装：
    pip install pandas openpyxl patchright

运行环境：
- Linux: 直接运行
- Windows: 需要修改 WORKSPACE 路径
"""

import pandas as pd
import sys
import logging
from datetime import datetime, timedelta
from pathlib import Path

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# 导入爬虫模块
from mohurd_scraper import MohurdScraper, query_mohurd_batch

# 配置
WORKSPACE = Path("/root/.openclaw/workspace-supplier")  # Linux 路径
CHUNK_SIZE = 500  # 每批处理 500 条

# Windows 用户请使用以下路径（需手动修改）
# WORKSPACE = Path(r"C:\Users\Administrator\.openclaw\workspace-supplier")


class SupplierUpdater:
    """供应商信息自动更新器"""
    
    def __init__(self, workspace_path=None):
        self.filepath = None
        self.updates = []
        self.workspace = Path(workspace_path) if workspace_path else WORKSPACE
        
        # Excel 列索引配置 (0-based)
        self.company_col_idx = 3      # 供应商列
        self.safety_cert_col_idx = 20  # 安全许可证有效期列
        
        # 缓存查询结果，避免重复查询
        self.query_cache = {}
        
        # 爬虫实例（复用）
        self.scraper = None
        
    def _init_scraper(self):
        """初始化爬虫"""
        if self.scraper is None:
            self.scraper = MohurdScraper(headless=True)
            self.scraper.start()
            logger.info("爬虫已启动")
    
    def _close_scraper(self):
        """关闭爬虫"""
        if self.scraper:
            self.scraper.close()
            self.scraper = None
            logger.info("爬虫已关闭")
        
    def _find_excel_file(self):
        """查找目标 Excel 文件"""
        if not self.workspace.exists():
            raise FileNotFoundError(f"工作目录不存在: {self.workspace}")
        
        files = list(self.workspace.glob("*.xlsx"))
        files = [f for f in files if '_backup' not in f.name and '_updated' not in f.name and '合格' in f.name]
        
        if not files:
            files = list(self.workspace.glob("*.xlsx"))
            files = [f for f in files if '_backup' not in f.name and '_updated' not in f.name]
        
        if not files:
            raise FileNotFoundError(f"未找到 Excel 文件: {self.workspace}")
        
        return files[0]
    
    def _get_column_names(self, filepath):
        """快速获取列名"""
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        wb.close()
        return headers
    
    def load_expiring_suppliers(self, days=30):
        """流式加载即将到期的供应商"""
        import openpyxl
        
        self.filepath = self._find_excel_file()
        logger.info(f"[1] 扫描文件: {self.filepath.name}")
        
        # 获取列名
        headers = self._get_column_names(self.filepath)
        total_rows = sum(1 for _ in openpyxl.load_workbook(self.filepath, read_only=True).active.iter_rows()) - 1
        
        self.company_col = headers[self.company_col_idx] if self.company_col_idx < len(headers) else None
        self.safety_cert_col = headers[self.safety_cert_col_idx] if self.safety_cert_col_idx < len(headers) else None
        
        logger.info(f"    总行数: {total_rows}")
        logger.info(f"    公司列: '{self.company_col}'")
        logger.info(f"    安全证有效期列: '{self.safety_cert_col}'")
        
        if not self.safety_cert_col:
            logger.warning("未找到有效期列")
            return pd.DataFrame()
        
        # 计算筛选日期
        today = datetime.now()
        cutoff = today + timedelta(days=days)
        
        logger.info(f"\n[2] 筛选 {days} 天内到期的记录 (截止日期: {cutoff.strftime('%Y-%m-%d')})")
        
        expiring_list = []
        
        # 流式读取
        wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
        ws = wb.active
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            company = row[self.company_col_idx] if self.company_col_idx < len(row) else None
            cert_date_raw = row[self.safety_cert_col_idx] if self.safety_cert_col_idx < len(row) else None
            
            if cert_date_raw:
                try:
                    if isinstance(cert_date_raw, datetime):
                        cert_date = cert_date_raw
                    elif isinstance(cert_date_raw, str):
                        cert_date = datetime.strptime(cert_date_raw.split()[0], '%Y-%m-%d')
                    else:
                        continue
                    
                    if cert_date <= cutoff:
                        expiring_list.append({
                            'row_idx': row_idx,
                            'company': company,
                            'cert_date': cert_date
                        })
                except (ValueError, TypeError):
                    continue
            
            if row_idx % 1000 == 0:
                logger.info(f"    已扫描 {row_idx}/{total_rows} 行，找到 {len(expiring_list)} 条待处理")
        
        wb.close()
        
        logger.info(f"\n    ✓ 扫描完成: {len(expiring_list)} 条记录需要处理")
        
        return pd.DataFrame(expiring_list)
    
    def query_mohurd_batch(self, companies):
        """
        批量查询住建部官网
        
        Args:
            companies: 公司名称列表
            
        Returns:
            字典 {公司名: {'date': 日期, 'source': '住建部官网'}}
        """
        # 先检查缓存
        to_query = [c for c in companies if c not in self.query_cache and c not in (None, '')]
        
        if not to_query:
            logger.info("[住建部] 全部命中缓存，跳过查询")
            return {c: self.query_cache[c] for c in companies if c in self.query_cache}
        
        logger.info(f"[住建部] 批量查询 {len(to_query)} 家公司...")
        
        try:
            # 初始化爬虫
            self._init_scraper()
            
            # 批量查询
            results = self.scraper.batch_query(to_query, timeout=60)
            
            # 更新缓存
            self.query_cache.update(results)
            
            logger.info(f"[住建部] 查询完成，命中 {len(results)} 条")
            return results
            
        except Exception as e:
            logger.error(f"[住建部] 查询异常: {e}")
            return {}
    
    def query_biaozhaozhao_batch(self, companies):
        """
        批量查询标招网（暂未接入）
        
        注意：标招网需要会员账号，暂不接入
        """
        logger.info("[标招网] 暂未接入（需要会员账号），跳过")
        return {}
    
    def apply_updates(self, expiring_df, mohurd_results):
        """应用更新到原始文件"""
        import openpyxl
        
        if expiring_df.empty:
            return 0
        
        wb = openpyxl.load_workbook(self.filepath)
        ws = wb.active
        
        updates_applied = 0
        
        for _, row in expiring_df.iterrows():
            row_idx = row['row_idx']
            company = row['company']
            
            # 只处理住建部结果
            if company in mohurd_results:
                result = mohurd_results[company]
                old_val = ws.cell(row=row_idx, column=self.safety_cert_col_idx + 1).value
                ws.cell(row=row_idx, column=self.safety_cert_col_idx + 1).value = result['date']
                
                self.updates.append({
                    'row': row_idx,
                    'company': str(company),
                    'cert_type': 'safety',
                    'old_date': str(old_val) if old_val else '未知',
                    'new_date': result['date'],
                    'source': result['source']
                })
                updates_applied += 1
                logger.info(f"    ✓ {company}: {old_val} -> {result['date']} (住建部)")
        
        # 保存备份
        backup_path = str(self.filepath).replace('.xlsx', '_backup.xlsx')
        wb.save(backup_path)
        logger.info(f"\n[4] 备份已保存: {backup_path}")
        
        # 保存更新文件
        updated_path = str(self.filepath).replace('.xlsx', '_updated.xlsx')
        wb.save(updated_path)
        logger.info(f"    更新文件已保存: {updated_path}")
        
        wb.close()
        return updates_applied
    
    def send_notification(self, to_email=None):
        """发送邮件通知"""
        from email_sender import send_supplier_update_report
        
        logger.info("\n[5] 发送邮件通知...")
        
        if not self.updates:
            logger.info("无更新内容，跳过邮件发送")
            return False
        
        try:
            report_lines = ["供应商信息更新报告", "=" * 40]
            report_lines.append(f"更新时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
            report_lines.append(f"更新数量: {len(self.updates)} 条")
            report_lines.append("")
            report_lines.append("更新详情:")
            for u in self.updates:
                report_lines.append(f"  - {u['company']}: {u['old_date']} -> {u['new_date']} (安全施工许可证, {u['source']})")
            
            report_text = "\n".join(report_lines)
            report_extra = self.filepath.name.replace('.xlsx', '_updated.xlsx') if self.filepath else ''
            
            success = send_supplier_update_report(self.updates, report_text, to_email, report_extra)
            
            if success:
                logger.info("    ✓ 邮件已发送")
            else:
                logger.error("    ✗ 邮件发送失败")
            return success
            
        except FileNotFoundError as e:
            logger.warning(f"邮件配置未找到: {e}")
            return False
        except Exception as e:
            logger.error(f"邮件发送异常: {e}")
            return False
    
    def run(self, days=30, send_email=True, to_email=None):
        """执行完整工作流"""
        logger.info("=" * 60)
        logger.info("🚀 供应商信息自动更新工作流 v5.0")
        logger.info("🏗️ 真实爬虫 +住建部官网 + Playwright")
        logger.info("=" * 60)
        
        try:
            # 1. 流式加载即将到期的数据
            logger.info("\n【步骤 1】流式扫描并筛选即将到期的供应商")
            expiring_df = self.load_expiring_suppliers(days)
            
            if expiring_df.empty:
                logger.info("✅ 无即将到期的供应商")
                return self.updates
            
            # 2. 批量查询
            logger.info("\n【步骤 2】批量查询数据源")
            companies = expiring_df['company'].dropna().unique().tolist()
            
            mohurd_results = self.query_mohurd_batch(companies)
            biao_results = self.query_biaozhaozhao_batch(companies)
            
            # 3. 应用更新
            logger.info("\n【步骤 3】应用更新到文件")
            updates_applied = self.apply_updates(expiring_df, mohurd_results)
            
            # 4. 邮件通知
            if send_email:
                self.send_notification(to_email)
            
            # 5. 报告
            logger.info("\n【步骤 6】生成报告")
            logger.info("=" * 60)
            logger.info(f"✅ 完成! 共更新 {len(self.updates)} 条记录")
            logger.info(f"📊 查询缓存: {len(self.query_cache)} 家公司")
            logger.info("=" * 60)
            
            return self.updates
            
        finally:
            # 确保关闭爬虫
            self._close_scraper()


def main(workspace_path=None):
    updater = SupplierUpdater(workspace_path)
    updates = updater.run(days=30, send_email=True)
    return updates


if __name__ == "__main__":
    import sys
    workspace_path = sys.argv[1] if len(sys.argv) > 1 else None
    updates = main(workspace_path)
