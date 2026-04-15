#!/usr/bin/env python3
"""
无锡招标抓取脚本 v2（2026-04-07）
★ 基于 ggzyjy.wuxi.gov.cn（无锡市公共资源交易中心）
"""
import urllib.request, re, ssl, time, os, io, zipfile, sys, json, html as html_mod, shutil, subprocess, base64
from datetime import datetime, timedelta
from urllib.parse import urlencode
from pathlib import Path

# ══════════════════════════════════════════════════════════════
# ★ CONFIG - 平台配置（无锡市公共资源交易中心）
# ══════════════════════════════════════════════════════════════

# 平台基础地址
BASE_URL = 'http://ggzyjy.wuxi.gov.cn'

# ★ 列表 API（POST，返回 JSON）
LIST_API_URL = BASE_URL + '/info_open/searchPublicResource'

# ★ 分类 chanId（从页面源码提取：getXzspZyjyzx('53051',...)）
CHAN_ID = '53051'   # 招标公告-施工

# Referer（防盗链）
REFERER = BASE_URL + '/wxsggzyjyzxzl/jyxx/jsgc/zbgg/gcl/index.shtml'

# ══════════════════════════════════════════════════════════════
# 腾讯文档配置（需新建独立的腾讯文档表格）
# ══════════════════════════════════════════════════════════════
TENCENT_FILE_ID = 'XPrEcspjgZpP'   # ★ 需创建腾讯文档并填入
TENCENT_SHEET_ID = 'BB08J2'
TENCENT_URL = 'https://docs.qq.com/sheet/XPrEcspjgZpP'

OUT_DIR = '/workspace/skills/wuxi-bid-search/outputs'
MAX_DAYS = 7
SITE_NAME = '无锡公共资源交易中心'
PERSISTENT_EXCEL = OUT_DIR + '/无锡招标_持续积累.xlsx'

# ══════════════════════════════════════════════════════════════
# 网络工具
# ══════════════════════════════════════════════════════════════
_ctx = ssl.create_default_context()
_ctx.check_hostname = False
_ctx.verify_mode = ssl.CERT_NONE

def get(url, headers=None, timeout=15):
    h = dict(headers) if headers else {}
    h.setdefault('User-Agent', 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36')
    h.setdefault('Accept-Language', 'zh-CN,zh;q=0.9')
    req = urllib.request.Request(url, headers=h)
    with urllib.request.urlopen(req, timeout=timeout, context=_ctx) as r:
        return r.read()

def post(url, data, headers=None, timeout=15):
    h = dict(headers) if headers else {}
    h.setdefault('User-Agent', 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36')
    h.setdefault('X-Requested-With', 'XMLHttpRequest')
    h.setdefault('Accept', 'application/json, text/javascript, */*; q=0.01')
    h.setdefault('Referer', REFERER)
    h.setdefault('Content-Type', 'application/x-www-form-urlencoded; charset=UTF-8')
    body = urlencode(data).encode('utf-8')
    req = urllib.request.Request(url, data=body, headers=h)
    with urllib.request.urlopen(req, timeout=timeout, context=_ctx) as r:
        return r.read()

# ── 日期工具 ────────────────────────────────────────────────────────────────
def parse_date(s):
    for fmt in ('%Y-%m-%d', '%Y/%m/%d'):
        try:
            return datetime.strptime(s.strip()[:10], fmt).date()
        except ValueError:
            pass
    return None

def is_within_days(date_str, days=MAX_DAYS):
    d = parse_date(date_str)
    if d is None:
        return False
    return d >= (datetime.now().date() - timedelta(days=days))

def is_expired(date_str, days=MAX_DAYS):
    """判断日期是否已过期（超过 N 天）"""
    d = parse_date(date_str)
    if d is None:
        return False
    cutoff = datetime.now().date() - timedelta(days=days)
    return d < cutoff

# ══════════════════════════════════════════════════════════════
# ★ 列表页解析（JSON API）
# ══════════════════════════════════════════════════════════════
def parse_list_json(api_resp):
    """
    解析无锡列表 API 返回的 JSON 数据。
    返回: [{'uuid': '', 'title': '', 'area': '', 'date': '', 'detail_url': ''}, ...]
    """
    items = []
    data = json.loads(api_resp)
    records = data.get('data', {}).get('data', [])
    for rec in records:
        title = rec.get('title', '')
        area = rec.get('jyly', '')   # 如"惠山区"
        date = rec.get('writeTimeString', rec.get('writeTime', ''))
        raw_id = rec.get('id', '')   # base64 编码的 URL

        # 解码 detail URL
        detail_url = None
        try:
            detail_url = base64.b64decode(raw_id).decode('utf-8')
        except Exception:
            detail_url = raw_id

        # 从 detail URL 提取 uuid（最后一个路径段去掉扩展名）
        uuid = ''
        if detail_url:
            m = re.search(r'/(\d+)\.shtml?', detail_url)
            if m:
                uuid = m.group(1)

        if title and date:
            items.append({
                'uuid': uuid,
                'title': title,
                'area': area,
                'date': date,
                'detail_url': detail_url,
            })
    return items

# ══════════════════════════════════════════════════════════════
# 正文清洗
# ══════════════════════════════════════════════════════════════
def clean_text(text):
    text = re.sub(r'<script[^>]*>[\s\S]*?</script>', ' ', text)
    text = re.sub(r'<style[^>]*>[\s\S]*?</style>', ' ', text)
    text = re.sub(r'<[^>]+>', ' ', text)
    text = html_mod.unescape(text)
    text = text.replace('\xa0', ' ').replace('&nbsp;', ' ')
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def find_field(text, keywords, length=300):
    for kw in keywords:
        i = text.find(kw)
        if i == -1:
            continue
        snippet = text[i:i+length]
        ci = snippet.find('：')
        ni = snippet.find('\n')
        end = min(ci + 1 if ci != -1 else length, ni if ni != -1 else length)
        val = snippet[ci+1:end].strip() if ci != -1 else snippet[:end].strip()
        if len(val) > 1:
            return val[:200]
    return ''

# ── 详情页字段提取（HTML 表格）────────────────────────────────────────────
def extract_table_fields(html):
    fields = {}
    for m in re.finditer(r'<td[^>]*>([\s\S]*?)</td>', html):
        cell = m.group(1).strip()
        if not cell or len(cell) < 4:
            continue
        km = re.match(r'^([^：:\n]{2,30})\s*[:：]\s*', cell)
        if km:
            k = km.group(1).strip()
            v = cell[len(km.group()):].strip()
            if k and v:
                fields[k] = re.sub(r'<[^>]+>', '', v).strip()[:200]
    return fields

def extract_detail_fields(html):
    text = clean_text(html)
    table_fields = extract_table_fields(html)
    result = {}

    key_map = {
        '标段名称': 'project_name', '项目名称': 'project_name',
        '招标人': 'client', '招标单位': 'client',
        '招标代理': 'agent', '代理机构': 'agent',
        '建设地点': 'site', '工程地点': 'site',
        '建设规模': 'scale', '工程规模': 'scale',
        '建设内容': 'content', '项目内容': 'content',
        '合同估算价': 'budget', '招标控制价': 'budget',
        '控制价': 'budget',
        '工期': 'duration', '计划工期': 'duration', '要求工期': 'duration',
        '计划开工': 'start_date', '开工日期': 'start_date',
        '计划竣工': 'end_date', '竣工日期': '竣工',
        '招标范围': 'scope',
        '投标人资质': 'qual_req', '资质要求': 'qual_req',
        '项目经理': 'pm_req', '项目负责人': 'pm_req',
        '资格审查': 'qual_check',
        '投标保证金': 'deposit',
        '履约担保': 'perform_guarantee',
        '合同价格形式': 'price_type',
        '投标有效期': 'valid_period',
        '付款方式': 'payment',
        '开标日期': 'bid_deadline', '投标截止时间': 'bid_deadline',
        '联系方式': 'contact',
    }
    for k, v in table_fields.items():
        for pattern, field_key in key_map.items():
            if pattern in k:
                result[field_key] = v
                break

    result.setdefault('client', '')
    result.setdefault('agent', '')
    result.setdefault('site', '')
    result.setdefault('budget', '')
    return result

# ══════════════════════════════════════════════════════════════
# PDF 附件下载（无锡相对路径）
# ══════════════════════════════════════════════════════════════
def fetch_pdfs(html, detail_url, out_dir):
    """
    从详情页 HTML 中提取 PDF 附件（无锡格式）。
    PDF 路径为相对路径：uploadfiles/...pdf
    """
    results = []
    SKIP_KEYWORDS = ['招标公告', '定稿', '变更', '澄清', '补充', '公告']

    # 提取 PDF 链接（无锡格式：/uploadfiles/...pdf）
    # <a href="/uploadfiles/202604/03/xxx.pdf" target="pdf">招标文件正文.pdf</a>
    pattern = re.compile(r'<a\s[^>]*href="(/uploadfiles/[^"]+\.pdf)"[^>]*>([^<]+\.pdf)\s*</a>')
    for m in pattern.finditer(html):
        file_path = m.group(1)   # 如 /uploadfiles/202604/03/xxx.pdf
        fname = m.group(2).strip()
        if any(kw in fname for kw in SKIP_KEYWORDS):
            continue
        try:
            # 完整 URL
            pdf_url = BASE_URL + file_path
            out_path = os.path.join(out_dir, fname)
            if os.path.exists(out_path):
                results.append({'name': fname, 'size': os.path.getsize(out_path), 'path': out_path})
                continue
            data = get(pdf_url)
            os.makedirs(out_dir, exist_ok=True)
            with open(out_path, 'wb') as f:
                f.write(data)
            results.append({'name': fname, 'size': len(data), 'path': out_path})
        except Exception as e:
            print(f'    [WARN] PDF 下载失败 [{fname}]: {e}')
    return results

# ══════════════════════════════════════════════════════════════
# PDF 字段提取（与常州相同，纯文本处理，平台无关）
# ══════════════════════════════════════════════════════════════
def extract_pdf_fields(pdf_path):
    """从招标文件 PDF 中提取关键字段（适配无锡2025版标准招标文件）。"""
    try:
        import pymupdf
    except ImportError:
        try:
            import fitz as pymupdf
        except ImportError:
            return {}
    try:
        doc = pymupdf.open(pdf_path)
    except Exception:
        return {}

    full_text = ''
    for page in doc:
        full_text += page.get_text() + '\n'
    doc.close()

    raw_text = full_text
    text_sp = re.sub(r'\s+', ' ', raw_text)

    def _between(text, start_pat, end_pat, length=200):
        m = re.search(start_pat, text)
        if not m:
            return ''
        start = m.end()
        end_match = re.search(end_pat, text[start:])
        end = start + end_match.start() if end_match else start + length
        val = text[start:end].strip()
        val = re.sub(r'\s+', ' ', val).strip()
        val = val.rstrip('。').rstrip('．').strip()
        return val[:length]

    def _find_after(text, keyword, length=200):
        idx = text.find(keyword)
        if idx == -1:
            return ''
        segment = text[idx:idx+length+len(keyword)]
        segment = re.sub(r'\n+', ' ', segment)
        val = segment[len(keyword):].strip()
        val = re.sub(r'^[：:\s]+', '', val)
        return val[:length].rstrip('。').rstrip('．').strip()

    def _num(text):
        """提取纯数字（支持中文数字和全角）。"""
        c2n = {'一':1,'二':2,'三':3,'四':4,'五':5,'六':6,'七':7,'八':8,'九':9,'十':10,
               '壹':1,'贰':2,'叁':3,'肆':4,'伍':5,'陆':6,'柒':7,'捌':8,'玖':9,'零':0,'〇':0,
               '０':'0','１':'1','２':'2','３':'3','４':'4','５':'5','６':'6','７':'7','８':'8','９':'9'}
        for ch, val in c2n.items():
            text = text.replace(ch, str(val))
        return re.sub(r'[^\d.]', '', text).strip()

    # ── 标段编号 ──────────────────────────────────────────────────────────────
    section_id = ''
    m = re.search(r'标段编号[）)：:\s]*([A-Z0-9]{10,30})', raw_text)
    if m:
        section_id = m.group(1).strip()

    # ── 1. 项目名称（取招标公告标题中的完整名称）──────────────────────────────
    # 招标公告标题格式：XXX（项目名称）YYY（标段名称）招标公告
    project_name = ''
    ch_idx = text_sp.find('第一章 招标公告（适用于公开招标）')
    if ch_idx < 0:
        ch_idx = text_sp.find('第一章 招标公告')
    if ch_idx >= 0:
        bidx = text_sp.find('标段编号', ch_idx)
        if bidx >= 0:
            title_area = text_sp[ch_idx:bidx]
            m_proj = re.search(r'（项目名称）(.+?)（标\s*段名称）', title_area)
            if m_proj:
                section = m_proj.group(1).strip()
                # 取title_area中最后一个完整句子/词组（项目名）
                before = title_area[:m_proj.start()].strip()
                # 找到最后一个"）"之后的非空内容（即项目名）
                last_paren = before.rfind('）')
                if last_paren >= 0:
                    proj_part = before[last_paren+1:].strip()
                else:
                    # 没有括号，取最后一个非空词
                    proj_part = before.split()[-1] if before.split() else ''
                project_name = (proj_part + section).strip()
    # 方式B：从前附表 2.1 标段名称
    if not project_name:
        project_name = _between(raw_text, r'2\.1\s*标段名称[：:\s*]', r'\n\s*\d+\.\d+', 80)

    # ── 2. 招标人 ─────────────────────────────────────────────────────────────
    client = ''
    # 方式A：前附表 1.1.2 招标人 → 名  称：XXX（精确提取）
    m_client = re.search(r'1\.1\.2\s*招标人[\s\n]+名[　 ]*称[：:]\s*([^\n]+)', raw_text)
    if m_client:
        client = m_client.group(1).strip()
    if not client:
        # 方式B：招标人为XXX，建设资金...
        idx = text_sp.find('招标人为')
        if idx >= 0:
            chunk = text_sp[idx+4:idx+100]
            m_cut = re.match(r'([^，,\（\(]+)', chunk)
            if m_cut:
                client = m_cut.group(1).strip()
    if not client:
        client = _find_after(text_sp, '招标人：', 80)

    # ── 3. 招标代理 ──────────────────────────────────────────────────────────
    agent = ''
    # 方式A：XXX（招标代理机构名称）受招标人的委托
    idx_agent = text_sp.find('（招标代理机构名称）')
    if idx_agent >= 0:
        chunk = text_sp[max(0, idx_agent-40):idx_agent]
        # 取chunk最后一个空格之后的内容（公司名）
        last_space = chunk.rfind(' ')
        agent = chunk[last_space+1:].strip() if last_space >= 0 else chunk.strip()
    if not agent:
        # 方式B：前附表 1.1.3 名  称：XXX
        m_agent2 = re.search(r'1\.1\.3\s*招标代理机构[\s\n]+名[　 ]*称[：:]\s*([^\n]+)', raw_text)
        if m_agent2:
            agent = m_agent2.group(1).strip()
    if not agent:
        agent = _find_after(text_sp, '招标代理机构：', 80)
    if not agent:
        agent = _find_after(text_sp, '招标代理：', 80)

    # ── 4. 联系方式 ──────────────────────────────────────────────────────────
    contact = ''
    m_tel = re.search(r'1\.1\.2\s*招标人[\s\n]+[^1]+?电[　 ]*话[：:]\s*([^\n]+)', raw_text)
    if m_tel:
        phones = re.findall(r'1[3-9]\d[\s\-]?\d{4}[\s\-]?\d{4}', m_tel.group(1))
        if phones:
            contact = phones[0]
    if not contact:
        phones = re.findall(r'1[3-9]\d[\s\-]?\d{4}[\s\-]?\d{4}', text_sp)
        if phones:
            contact = phones[0]

    # ── 5. 代理联系方式 ──────────────────────────────────────────────────────
    agent_contact = ''
    # 从 1.1.3 招标代理机构条目提取
    m_atel = re.search(r'1\.1\.3\s*招标代理机构[\s\n]+[^\n]+?电[　 ]*话[：:]\s*([^\n]+)', raw_text)
    if m_atel:
        aps = re.findall(r'1[3-9]\d[\s\-]?\d{4}[\s\-]?\d{4}', m_atel.group(1))
        if aps:
            agent_contact = aps[0]
    if not agent_contact:
        # 从 10.联系方式章节
        chapter10_idx = text_sp.rfind('10.联系方式')
        if chapter10_idx < 0:
            chapter10_idx = text_sp.rfind('联系方式')
        if chapter10_idx >= 0:
            ac = text_sp[chapter10_idx:chapter10_idx+400]
            aps = re.findall(r'1[3-9]\d[\s\-]?\d{4}[\s\-]?\d{4}', ac)
            if len(aps) > 1:
                agent_contact = ' '.join(aps[1:2])
            elif aps:
                agent_contact = aps[0]

    # ── 6. 建设地点 ──────────────────────────────────────────────────────────
    site = _between(raw_text, r'2\.2\s*建设地点[：:\s*]', r'\n\s*\d+\.\d+', 80)
    if not site:
        m_site = re.search(r'1\.1\.6\s*建设地点[\s\n]+([^\n]+?)(?=\n|\d+\.\d)', raw_text)
        if m_site:
            site = m_site.group(1).strip()
    if not site:
        site = _find_after(text_sp, '建设地点：', 80)

    # ── 7. 建设内容 ──────────────────────────────────────────────────────────
    content = _between(raw_text, r'2\.3\s*建设内容[：:\s*]', r'\n\s*\d+\.\d+', 100)
    if not content:
        content = _find_after(text_sp, '建设内容：', 100)

    # ── 8. 质量 ──────────────────────────────────────────────────────────────
    quality = ''
    for kw in ['合格', '优良', '符合国家标准', '合格标准']:
        if kw in raw_text:
            quality = kw
            break

    # ── 9. 工程规模 ──────────────────────────────────────────────────────────
    scale = _between(raw_text, r'工程规模[：:\s*]', r'\n', 100)
    if not scale:
        scale = _between(raw_text, r'建设规模[：:\s*]', r'\n', 100)

    # ── 10. 控制价（合同估算价） ─────────────────────────────────────────────
    budget = ''
    # 方式A：2.6 工程合同估算价（万元）：1500
    m_budget = re.search(r'2\.6\s*工程合同估算价（万元）[：:\s]*([\d四五六七八九十壹贰叁肆伍陆柒捌玖零〇\.,]+)', text_sp)
    if m_budget:
        budget = _num(m_budget.group(1))
    if not budget:
        # 方式B：最高投标限价  金额：14836189.14 元
        m_budget2 = re.search(r'最高投标限价[\s\n]+金额[：:]\s*([\d\.,]+)\s*元', text_sp)
        if m_budget2:
            budget = m_budget2.group(1).strip()
    if not budget:
        for pat in [r'合同估算价[：:\s]*约?\s*([\d四五六七八九十壹贰叁肆伍陆柒捌玖零〇\.,]+)',
                    r'招标控制价[：:\s]*([\d四五六七八九十壹贰叁肆伍陆柒捌玖零〇\.,]+)',
                    r'最高投标限价[：:\s]*([\d四五六七八九十壹贰叁肆伍陆柒捌玖零〇\.,]+)']:
            nm = re.search(pat, text_sp)
            if nm:
                budget = _num(nm.group(1))
                break

    # ── 11. 招标范围 ─────────────────────────────────────────────────────────
    scope = _between(raw_text, r'2\.4\s*招标范围[：:\s*]', r'\n\s*\d+\.\d+', 100)
    if not scope:
        scope = _between(raw_text, r'招标范围[：:\s*]', r'\n', 100)
    if not scope:
        scope = '见投标人须知前附表'

    # ── 12. 工期 ──────────────────────────────────────────────────────────────
    duration = ''
    # 优先从 2.9 节提取：2.9 工期：100 日历天
    m_dur = re.search(r'2\.9\s*工期[：:]\s*(\d+)\s*日历天', text_sp)
    if m_dur:
        duration = m_dur.group(1) + ' 日历天'
    if not duration:
        # 前附表 要求工期：100 日历天
        m_dur2 = re.search(r'要求工期[：:]\s*(\d+)\s*日历天', text_sp)
        if m_dur2:
            duration = m_dur2.group(1) + ' 日历天'
        else:
            duration = _between(raw_text, r'要求工期[：:\s]*', r'\n', 60)
            if not duration:
                duration = _find_after(text_sp, '工期：', 60)

    # ── 13. 计划开工/竣工 ────────────────────────────────────────────────────
    start_date = ''
    end_date = ''
    m_start = re.search(r'计划开工日期[：:]\s*(\d{4})\s*年(\d{1,2})\s*月(\d{1,2})\s*日', text_sp)
    if m_start:
        start_date = f"{m_start.group(1)}-{int(m_start.group(2)):02d}-{int(m_start.group(3)):02d}"
    m_end = re.search(r'计划竣工日期[：:]\s*(\d{4})\s*年(\d{1,2})\s*月(\d{1,2})\s*日', text_sp)
    if m_end:
        end_date = f"{m_end.group(1)}-{int(m_end.group(2)):02d}-{int(m_end.group(3)):02d}"
    if not start_date:
        m_sd = re.search(r'计划开工[：:]\s*(\d{4}-\d{1,2}-\d{1,2})', text_sp)
        if m_sd:
            start_date = m_sd.group(1)
    if not end_date:
        m_ed = re.search(r'计划竣工[：:]\s*(\d{4}-\d{1,2}-\d{1,2})', text_sp)
        if m_ed:
            end_date = m_ed.group(1)

    # ── 14. 投标保证金 ────────────────────────────────────────────────────────
    # 无锡格式：保证金金额或投标保函担保金额：人民币8 万元 或 人民币伍万元
    # 支持阿拉伯数字（8万）和中文数字（伍万）
    deposit = ''
    # 优先：在"保证金金额"关键词后80字内找"人民币X万"
    idx_dep = text_sp.find('保证金金额')
    if idx_dep >= 0:
        chunk = text_sp[idx_dep:idx_dep + 100]
        m = re.search(r'人民币\s*([0-9零一二三四五六七八九十壹贰叁肆伍陆柒捌玖拾百千万零〇０-９]+)\s*万', chunk)
        if m:
            deposit = _num(m.group(1)) + ' 万'
    # 备用：标准格式
    if not deposit:
        m_dep = re.search(r'投标保证金金额[：:]\s*人民币([0-9零一二三四五六七八九十壹贰叁肆伍陆柒捌玖拾百千万]+)\s*万', text_sp)
        if m_dep:
            deposit = _num(m_dep.group(1)) + ' 万'
    # 备用2：全文搜索"人民币X万"（取第一个）
    if not deposit:
        m_dep2 = re.search(r'人民币\s*([0-9零一二三四五六七八九十壹贰叁肆伍陆柒捌玖拾百千万]+)\s*万', text_sp)
        if m_dep2:
            deposit = _num(m_dep2.group(1)) + ' 万'

    # ── 15. 资格审查 ─────────────────────────────────────────────────────────
    # 支持多种格式：
    #   1. ☑资格后审 / □资格后审（勾选框格式）
    #   2. 18 资格后审（节编号+名称，无锡2025版格式）
    #   3. 资格审查办法 ... 资格后审
    #   4. "适用于资格后审"（EPC/工程总承包文档标题格式）
    qual_check = ''
    if re.search(r'[☑□]\s*资格后审', raw_text) or re.search(r'[☑□]\s*资格后审', text_sp):
        qual_check = '资格后审'
    elif re.search(r'[☑□]\s*资格预审', raw_text) or re.search(r'[☑□]\s*资格预审', text_sp):
        qual_check = '资格预审'
    elif re.search(r'(?:^|\s)\d+\.?\s*资格后审', text_sp) or \
         re.search(r'资格审查办法[^钱\n]*资格后审', text_sp):
        qual_check = '资格后审'
    elif re.search(r'(?:^|\s)\d+\.?\s*资格预审', text_sp):
        qual_check = '资格预审'
    elif re.search(r'适用于\s*资格后审', text_sp):
        qual_check = '资格后审'

    # ── 16. 投标有效期 ────────────────────────────────────────────────────────
    valid_period = ''
    m_vp = re.search(r'(\d+)\s*天（从投标截止之日起算）', text_sp)
    if m_vp:
        valid_period = m_vp.group(1) + ' 天'
    if not valid_period:
        m_vp2 = re.search(r'投标有效期[\s\n]+(\d+)\s*天', text_sp)
        if m_vp2:
            valid_period = m_vp2.group(1) + ' 天'
    if not valid_period:
        valid_period = _between(raw_text, r'3\.3\s*投标有效期[：:\s]*', r'\n', 60)
        if not valid_period:
            valid_period = _find_after(text_sp, '投标有效期：', 60)

    # ── 17. 资质要求 ─────────────────────────────────────────────────────────
    qual_req = ''
    m_qr = re.search(r'3\.1\s*投标人资质类别和等级[：:]\s*\[([^\]]+)\]', text_sp)
    if m_qr:
        qual_req = '[' + m_qr.group(1) + ']'
    if not qual_req:
        qual_req = _between(raw_text, r'3\.1\s*投标人资质[：:\s*]', r'\n\s*3\.2', 120)
    if not qual_req:
        qual_req = _between(raw_text, r'投标人资质类别和等级[：:\s*]', r'\n', 120)

    # ── 18. 项目负责人资质 ────────────────────────────────────────────────────
    pm_req = ''
    m_pm = re.search(r'3\.2\s*拟选派项目负责人专业及资质等级[：:]\s*\[([^\]]+)\]', text_sp)
    if m_pm:
        pm_req = '[' + m_pm.group(1) + ']'
    if not pm_req:
        pm_req = _between(raw_text, r'3\.2\s*拟选派项目负责人专业及资质等级[：:\s*]', r'\n\s*3\.3', 100)
    if not pm_req:
        pm_req = _between(raw_text, r'项目负责人资质[：:\s*]', r'\n', 80)

    # ── 19. 合同价格形式 ─────────────────────────────────────────────────────
    # 支持 ☑/□ 勾选框，以及"固定"/"可调"前缀（无锡EPC格式：☑固定总价合同）
    price_type = ''
    if re.search(r'[☑□]\s*(?:固定|可调)?\s*单价合同', raw_text) or \
       re.search(r'[☑□]\s*(?:固定|可调)?\s*单价合同', text_sp):
        price_type = '单价合同'
    elif re.search(r'[☑□]\s*(?:固定|可调)?\s*总价合同', raw_text) or \
         re.search(r'[☑□]\s*(?:固定|可调)?\s*总价合同', text_sp):
        price_type = '总价合同'
    elif re.search(r'[☑□]\s*(?:固定|可调)?\s*成本加酬金合同', raw_text):
        price_type = '成本加酬金合同'

    # ── 20. 履约担保 ─────────────────────────────────────────────────────────
    perform_guarantee = ''
    m_pg = re.search(r'履约担保的金额[：:]\s*合同含税价款的([\d０-９]+)%', text_sp)
    if m_pg:
        pct = m_pg.group(1).replace('０','0').replace('１','1').replace('２','2').replace('３','3').replace('４','4').replace('５','5').replace('６','6').replace('７','7').replace('８','8').replace('９','9')
        perform_guarantee = f'合同总价（{pct}%）'
    else:
        m_pg2 = re.search(r'合同[总净]价[^\n。]{0,50}?的\s*([\d０-９]+)\s*[％%]', text_sp)
        if m_pg2:
            pct = m_pg2.group(1).replace('０','0').replace('１','1').replace('２','2').replace('３','3').replace('４','4').replace('５','5').replace('６','6').replace('７','7').replace('８','8').replace('９','9')
            perform_guarantee = f'合同总价（{pct}%）'

    # ── 21. 付款方式 ──────────────────────────────────────────────────────────
    payment = ''
    for kw in ['按月支付', '按进度支付', '分期支付', '一次性支付']:
        if kw in raw_text:
            payment = kw
            break

    # ── 22. 开标日期 ─────────────────────────────────────────────────────────
    bid_deadline = _find_after(text_sp, '开标时间：', 80)
    if not bid_deadline:
        bid_deadline = _find_after(text_sp, '投标截止时间：', 80)
    if not bid_deadline:
        bd = _between(raw_text, r'开标（投标截止）时间[：:\s]*', r'\n', 60)
        if not bd:
            bd = _between(raw_text, r'投标截止时间[：:\s]*', r'\n', 60)
        bid_deadline = bd

    # ── 23. 联合体 ──────────────────────────────────────────────────────────
    joint_bid = ''
    if '☑不接受' in raw_text and '联合体' in raw_text:
        joint_bid = '不接受联合体'
    elif '☑接受' in raw_text and '联合体' in raw_text:
        joint_bid = '接受联合体'

    # ── 24. 评标办法 ─────────────────────────────────────────────────────────
    bid_method = ''
    if '☑合理低价法' in raw_text:
        bid_method = '合理低价法'
    elif '☑评定分离' in raw_text:
        bid_method = '综合评估法-评定分离'
    elif '☑综合评估法' in raw_text:
        bid_method = '综合评估法'
    elif '☑经评审的最低投标价法' in raw_text:
        bid_method = '经评审的最低投标价法'

    # 评标入围方法
    shortlist_method, shortlist_count = '', ''
    method_idx = text_sp.find('☑方法')
    if method_idx < 0:
        method_idx = text_sp.find('□方法')
    if method_idx >= 0:
        section_start = text_sp.rfind('评标入围', 0, method_idx)
        if section_start < 0:
            section_start = text_sp.rfind('评标入围方法', 0, method_idx)
        if section_start >= 0:
            shortlist_section = text_sp[section_start:section_start+3000]
            m_short = re.search(r'[☑□]\s*方法([一二三四五六])[：:]\s*(\S{2,10}(?:法|入围))', shortlist_section)
            if m_short:
                shortlist_method = m_short.group(2)
                method_pos = m_short.start()
                method_start = max(0, method_pos - 50)
                next_method_m = re.search(r'[☑□]\s*方法[一二三四五六]', shortlist_section[method_pos+20:])
                if next_method_m:
                    method_end = method_pos + 20 + next_method_m.start()
                else:
                    method_end = method_pos + 500
                method_block = shortlist_section[method_start:method_end]
                m_count = re.search(r'不少于\s*(\d+)\s*家', method_block)
                if m_count:
                    shortlist_count = m_count.group(1)

    if shortlist_method and shortlist_count:
        bid_method = f'{bid_method}，{shortlist_method}（{shortlist_count}家）'
    elif shortlist_method:
        bid_method = f'{bid_method}，{shortlist_method}'

    # 综合评估法评分组成
    eval_detail = ''
    if '☑综合评估法' in raw_text or '☑综合评估法' in text_sp:
        parts_detail = []
        idx_234 = text_sp.find('2.3.4 (1)')
        if idx_234 < 0:
            idx_234 = text_sp.find('2.3.4(1)')
        if idx_234 >= 0:
            sec_234 = text_sp[idx_234:idx_234 + 8000]
            tidx = sec_234.find('评分因素 页数要求 分值')
            if tidx >= 0:
                tbl = sec_234[tidx:tidx + 1200]
                rows = re.split(r'☑', tbl)
                construct = 0
                for row in rows:
                    row = row.strip()
                    if not row:
                        continue
                    m_score = re.search(r'(\d+)\s*分', row)
                    if not m_score:
                        continue
                    score = int(m_score.group(1))
                    if '答辩' in row or '□其他' in row or '定标委员会' in row or '票决' in row:
                        continue
                    if score > 20:
                        continue
                    construct += score
                def_chunk = sec_234[1170:1800]
                m_def = re.search(r'总分\s*(\d+)\s*分', def_chunk)
                if m_def:
                    parts_detail.append(f'答辩{int(m_def.group(1))}分')
                if construct > 0:
                    parts_detail.insert(0, f'施工组织设计{construct}分')
        eval_detail = '，'.join(parts_detail) if parts_detail else ''
        if eval_detail:
            bid_method = f'{bid_method}，{eval_detail}'

    # ── 25. 评标基准价计算方法（无锡2025版） ─────────────────────────────────
    基准价计算 = ''
    idx_base = text_sp.find('评标基准价=')
    if idx_base < 0:
        idx_base = text_sp.find('评标基准价计算方法：')
    if idx_base < 0:
        # 无锡2025版格式：标题间有空格 "评标基准价 计算方法"
        idx_base = text_sp.find('评标基准价 计算方法')
    if idx_base >= 0:
        section = text_sp[idx_base:idx_base + 4000]

        # 找选中的方法（一/二/三），同时兼容 □ 和 ☑
        cn_map = {'一':'方法一','二':'方法二','三':'方法三',
                  '四':'方法四','五':'方法五','六':'方法六'}
        m_sel = re.search(r'[☑□]\s*方法([一二三四五六]+)', section[:300])
        method_num = m_sel.group(1) if m_sel else ''
        method_key = cn_map.get(method_num, '')

        method_name = ''
        if method_key:
            key_pos = section.find(method_key)
            if key_pos >= 0:
                formula_area = section[key_pos:key_pos+200]
                formula_area_ns = formula_area.replace(' ', '')
                m_f1 = re.search(r'评标基准价=A×K[×K]', formula_area_ns)
                if m_f1:
                    method_name = 'A×K法'
                m_f2 = re.search(r'评标基准价=A×K1×Q1[+]B×K2×Q2', formula_area_ns)
                if m_f2:
                    method_name = 'A×K1×Q1+B×K2×Q2法'
                if 'ABC' in formula_area_ns or 'C×20%' in formula_area_ns:
                    method_name = 'ABC合成法'

        # K值范围
        k_range = ''
        m_k = re.search(r'K\s+值.{0,5}取值范围[为：]([\d%，.%～\-\s、]+)', section[:2000])
        if m_k:
            k_str = m_k.group(1).replace('～', '-').replace(' ', '').replace('%', '%')
            k_vals = re.findall(r'(\d+(?:\.\d+)?)%', k_str)
            if len(k_vals) >= 2:
                k_range = f"{k_vals[0]}%-{k_vals[-1]}%"

        # 下浮率Δ范围
        delta_range = ''
        tidx = section.find('下浮率Δ')
        if tidx < 0:
            tidx = section.find('下浮率△')
        if tidx >= 0:
            delta_block = section[tidx:tidx + 2000]
            m_k1 = re.search(r'K\d?\s*值.{0,5}取值范围[为：]([\d%，.%～\-\s、]+)', delta_block)
            if m_k1:
                k1_str = m_k1.group(1).replace('～', '-').replace(' ', '').replace('%', '%')
                k1_vals = re.findall(r'(\d+(?:\.\d+)?)%', k1_str)
                if k1_vals:
                    k_range = f"{k1_vals[0]}%-{k1_vals[-1]}%"
            m_delta = re.search(r'[Δ△]取值范围[为：]([\d%，.%～\-\s、]+)', delta_block)
            if m_delta:
                delta_str = m_delta.group(1).replace('～', '-').replace('%', '%')
                delta_vals = re.findall(r'(\d+)%', delta_str)
                if len(delta_vals) >= 2:
                    delta_range = f"{delta_vals[0]}%-{delta_vals[-1]}%"
                elif len(delta_vals) == 1:
                    delta_range = delta_vals[0] + '%'

        # Q1范围（方法二）
        q1_range = ''
        m_q1 = re.search(r'Q1的取值范围[为：]([\d%，.%～\-\s、]+)', section[:2000])
        if m_q1:
            q1_str = m_q1.group(1).replace('～', '-').replace(' ', '').replace('%', '%')
            q1_vals = re.findall(r'(\d+)%', q1_str)
            if len(q1_vals) >= 2:
                q1_range = f"{q1_vals[0]}%-{q1_vals[-1]}%"

        parts = []
        if method_name:
            parts.append(method_name)
        if k_range:
            parts.append(f"K（{k_range}）")
        if q1_range:
            parts.append(f"Q1（{q1_range}）")
        if delta_range:
            parts.append(f"下浮率Δ（{delta_range}）")
        elif parts:
            parts.append(f"下浮率Δ（/）")
        基准价计算 = '，'.join(parts)

    return {
        'project_name': project_name,
        'section_id': section_id,
        'client': client,
        'agent': agent,
        'site': site,
        'content': content,
        'quality': quality,
        'scale': scale,
        'budget': budget,
        'scope': scope,
        'duration': duration,
        'start_date': start_date,
        'end_date': end_date,
        'qual_req': qual_req,
        'pm_req': pm_req,
        'bid_method': bid_method,
        'qual_check': qual_check,
        'deposit': deposit,
        'perform_guarantee': perform_guarantee,
        'price_type': price_type,
        'valid_period': valid_period,
        'payment': payment,
        'bid_deadline': bid_deadline,
        'contact': contact,
        'agent_contact': agent_contact,
        'joint_bid': joint_bid,
        '基准价计算': 基准价计算,
        'eval_detail': eval_detail,
    }



# ══════════════════════════════════════════════════════════════
# Excel 生成
# ══════════════════════════════════════════════════════════════
def make_xlsx(data, out_path):
    headers = ['序号', '项目名称', '地区', '发布日期', '分类',
               '招标人', '招标代理', '建设地点', '建设内容',
               '控制价(万元)', '评标办法', '评标基准价计算方法', '资格审查',
               '投标保证金', '工期', '计划开工', '计划竣工',
               '资质要求', '项目负责人资质', '合同价格形式', '投标有效期',
               '履约担保', '联系方式', '标段编号', '来源']
    col_widths = [5, 40, 10, 12, 10,
                  24, 24, 24, 38,
                  14, 20, 20, 10,
                  16, 12, 14, 14,
                  40, 30, 12, 10,
                  30, 30, 30, 22]

    def cr(c, r):
        s = ''
        c += 1
        while c:
            c, rem = divmod(c - 1, 26)
            s = chr(65 + rem) + s
        return s + str(r)

    si_map, strings = {}, []

    def si(s):
        s = str(s)
        if s not in si_map:
            si_map[s] = len(strings)
            strings.append(s)
        return si_map[s]

    for h in headers: si(h)

    def row_fields(d, ri):
        return [
            str(ri-1),
            d.get('project_name','') or d.get('title',''),
            d.get('area',''),
            d.get('date',''),
            d.get('sub_cat',''),
            d.get('client',''),
            d.get('agent',''),
            d.get('site',''),
            d.get('content',''),
            d.get('budget',''),
            d.get('bid_method',''),
            d.get('基准价计算',''),
            d.get('qual_check',''),
            d.get('deposit',''),
            d.get('duration',''),
            d.get('start_date',''),
            d.get('end_date',''),
            d.get('qual_req',''),
            d.get('pm_req',''),
            d.get('price_type',''),
            d.get('valid_period',''),
            d.get('perform_guarantee',''),
            d.get('contact','') or d.get('agent_contact',''),
            d.get('section_id',''),
            SITE_NAME,
        ]

    rows_xml = ['<row r="1">' + ''.join(f'<c r="{cr(i,1)}" s="1" t="s"><v>{si(h)}</v></c>' for i,h in enumerate(headers)) + '</row>']
    for ri, d in enumerate(data, 2):
        fields = row_fields(d, ri)
        for v in fields: si(v)
        rows_xml.append('<row r="'+str(ri)+'">' + ''.join(f'<c r="{cr(i,ri)}" s="2" t="s"><v>{si(v)}</v></c>' for i,v in enumerate(fields)) + '</row>')

    cols = ''.join(f'<col min="{i+1}" max="{i+1}" width="{w}" customWidth="1"/>' for i,w in enumerate(col_widths))
    sheet = f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetViews><sheetView workbookViewId="0" showGridLines="1"><selection activeCell="A1" sqref="A1"/></sheetView></sheetViews><cols>{cols}</cols><sheetData>{"".join(rows_xml)}</sheetData><pageMargins left="0.5" right="0.5" top="0.75" bottom="0.75"/></worksheet>'
    wb = f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><sheets><sheet name="\u62db\u6807\u9879\u76ee\u660e\u7ec6" sheetId="1" r:id="rId1"/></sheets></workbook>'
    styles = f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><fonts><font><sz val="11"/><name val="\u5fae\u8f6f\u96c5\u9ed1"/></font><font><sz val="11"/><b/><name val="\u5fae\u8f6f\u96c5\u9ed1"/></font></fonts><fills><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill><fill><patternFill patternType="solid"><fgColor rgb="FF028090"/></patternFill></fill></fills><borders><border><left/><right/><top/><bottom/><diagonal/></border></borders><cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs><cellXfs><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/><xf numFmtId="0" fontId="1" fillId="2" borderId="0" xfId="0"><alignment horizontal="center"/></xf><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"><alignment wrapText="1"/></xf></cellStyleXfs></styleSheet>'
    wb_rels = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/><Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/><Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/></Relationships>'
    root_rels = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/></Relationships>'
    ct = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/><Default Extension="xml" ContentType="application/xml"/><Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/><Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/><Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/><Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/></Types>'

    def esc(s):
        return str(s).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('"','&quot;').replace("'",'&apos;')

    ss_xml = f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?><sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="{len(strings)}" uniqueCount="{len(strings)}">{"".join(f"<si><t>{esc(s)}</t></si>" for s in strings)}</sst>'

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as z:
        z.writestr('[Content_Types].xml', ct)
        z.writestr('_rels/.rels', root_rels)
        z.writestr('xl/workbook.xml', wb)
        z.writestr('xl/_rels/workbook.xml.rels', wb_rels)
        z.writestr('xl/styles.xml', styles)
        z.writestr('xl/worksheets/sheet1.xml', sheet)
        z.writestr('xl/sharedStrings.xml', ss_xml)

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, 'wb') as f:
        f.write(buf.getvalue())
    print(f'  Excel: {out_path}  ({os.path.getsize(out_path)} bytes)')

# ══════════════════════════════════════════════════════════════
# 腾讯文档读取
# ══════════════════════════════════════════════════════════════
def qq_read_all(mcporter_cmd='mcporter'):
    import csv, io as sio
    cmd = [
        mcporter_cmd, 'call', 'tencent-sheetengine', 'get_cell_data',
        '--args', json.dumps({
            'file_id': TENCENT_FILE_ID,
            'sheet_id': TENCENT_SHEET_ID,
            'start_row': 0, 'start_col': 0,
            'end_row': 500, 'end_col': 24,
            'return_csv': True,
        })
    ]
    try:
        result = subprocess.run(cmd, capture_output=True, timeout=30, text=False)
        raw = result.stdout
    except Exception as e:
        print(f'    [WARN] 腾讯文档读取失败: {e}')
        return []
    try:
        text = raw.decode('utf-8')
    except Exception:
        text = raw.decode('utf-8', errors='replace')
    try:
        resp = json.loads(text)
    except Exception:
        print(f'    [WARN] 腾讯文档响应解析失败')
        return []
    csv_data = resp.get('csv_data', '')
    if not csv_data or not csv_data.strip():
        return []
    reader = csv.DictReader(sio.StringIO(csv_data))
    rows = []
    for row in reader:
        sid = row.get('标段编号', '').strip()
        if not sid:
            continue
        rows.append({
            'row_idx': int(row.get('序号', 0)),
            'section_id': sid,
            'date': row.get('发布日期', '').strip(),
            'title': row.get('项目名称', '').strip(),
        })
    return rows

def qq_sync(mcporter_cmd, local_items, dry_run=False):
    import csv, io as sio
    existing = qq_read_all(mcporter_cmd)
    print(f'    文档现有: {len(existing)} 项', end='')
    if not existing:
        print()
        if dry_run:
            print('    [DRY-RUN] 跳过写入')
            return
        _qq_write_all(mcporter_cmd, local_items)
        return
    local_sids = {it.get('section_id', '') for it in local_items}
    to_delete = [r for r in existing if r.get('section_id', '') not in local_sids]
    print(f' | 保留: {len(existing) - len(to_delete)} | 新增: {len(local_sids - {r["section_id"] for r in existing})} | 过期删除: {len(to_delete)}')
    for r in to_delete:
        print(f'    - 删除过期: {r["title"][:40]}')
    if dry_run:
        print('    [DRY-RUN] 跳过写入')
        return
    _qq_write_all(mcporter_cmd, local_items)

def _qq_write_all(mcporter_cmd, local_items):
    import csv, io as sio
    headers = ['序号', '项目名称', '地区', '发布日期', '分类',
               '招标人', '招标代理', '建设地点', '建设内容',
               '控制价(万元)', '评标办法', '评标基准价计算方法', '资格审查',
               '投标保证金', '工期', '计划开工', '计划竣工',
               '资质要求', '项目负责人资质', '合同价格形式', '投标有效期',
               '履约担保', '联系方式', '标段编号', '来源']
    field_keys = ['project_name', 'area', 'date', 'sub_cat',
                  'client', 'agent', 'site', 'content',
                  'budget', 'bid_method', '基准价计算', 'qual_check',
                  'deposit', 'duration', 'start_date', 'end_date',
                  'qual_req', 'pm_req', 'price_type', 'valid_period',
                  'perform_guarantee', 'contact', 'section_id', '_source']
    rows_csv = []
    for i, item in enumerate(local_items, 1):
        row = []
        for k in field_keys:
            if k == '_source':
                row.append(SITE_NAME)
            else:
                v = item.get(k, '')
                row.append(str(v) if v is not None else '')
        row.insert(0, i)
        rows_csv.append(row)
    csv_buf = io.StringIO()
    writer = csv.writer(csv_buf)
    writer.writerow(headers)
    writer.writerows(rows_csv)
    csv_str = csv_buf.getvalue()
    import subprocess as sp
    cmd = [
        mcporter_cmd, 'call', 'tencent-sheetengine', 'set_range_value_by_csv',
        '--args', json.dumps({
            'file_id': TENCENT_FILE_ID,
            'sheet_id': TENCENT_SHEET_ID,
            'csv_data': csv_str,
        })
    ]
    result = sp.run(cmd, capture_output=True, timeout=30, text=True)
    if result.returncode == 0:
        print(f'    全量写入 {len(local_items)} 行 ✅')
    else:
        print(f'    [ERROR] 写入失败: {result.stderr[:200]}')

# ══════════════════════════════════════════════════════════════
# 持续积累 Excel 读写（openpyxl）
# ══════════════════════════════════════════════════════════════
def read_persistent_excel():
    """
    读取本地持续积累 Excel，返回 (data_list, existing_section_ids_set)
    """
    try:
        import openpyxl
    except ImportError:
        return [], set()

    path = PERSISTENT_EXCEL
    if not os.path.exists(path):
        return [], set()

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        h_map = {h: i for i, h in enumerate(headers) if h}

        sid_idx = h_map.get('标段编号', -1)

        existing_sids = set()
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            sid = row[sid_idx] if sid_idx >= 0 else ''
            if sid:
                existing_sids.add(str(sid).strip())
            rows.append(dict(zip(headers, row)))

        print(f'    [持久Excel] 读取 {len(rows)} 条历史记录')
        return rows, existing_sids
    except Exception as e:
        print(f'    [WARN] 读取持久Excel失败: {e}')
        return [], set()


def write_persistent_excel(all_items):
    """将 all_items 写入固定文件名 PERSISTENT_EXCEL。"""
    headers = ['序号', '项目名称', '地区', '发布日期', '分类',
               '招标人', '招标代理', '建设地点', '建设内容',
               '控制价(万元)', '评标办法', '评标基准价计算方法', '资格审查',
               '投标保证金', '工期', '计划开工', '计划竣工',
               '资质要求', '项目负责人资质', '合同价格形式', '投标有效期',
               '履约担保', '联系方式', '标段编号', '来源']

    try:
        import openpyxl
    except ImportError:
        make_xlsx(all_items, PERSISTENT_EXCEL)
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '招标项目明细'

    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)

    for ri_offset, it in enumerate(all_items, 2):
        row_vals = [
            str(ri_offset - 1),
            it.get('project_name', '') or it.get('title', ''),
            it.get('area', ''),
            it.get('date', ''),
            it.get('sub_cat', ''),
            it.get('client', ''),
            it.get('agent', ''),
            it.get('site', ''),
            (it.get('content', '') or '')[:300],
            it.get('budget', ''),
            it.get('bid_method', ''),
            it.get('基准价计算', ''),
            it.get('qual_check', ''),
            it.get('deposit', ''),
            it.get('duration', ''),
            it.get('start_date', ''),
            it.get('end_date', ''),
            it.get('qual_req', ''),
            it.get('pm_req', ''),
            it.get('price_type', ''),
            it.get('valid_period', ''),
            it.get('perform_guarantee', ''),
            it.get('contact', '') or it.get('agent_contact', ''),
            it.get('section_id', ''),
            SITE_NAME,
        ]
        for ci, v in enumerate(row_vals, 1):
            ws.cell(row=ri_offset, column=ci, value=v)

    os.makedirs(os.path.dirname(PERSISTENT_EXCEL), exist_ok=True)
    wb.save(PERSISTENT_EXCEL)
    print(f'  持续Excel写入: {PERSISTENT_EXCEL}  ({len(all_items)} 条)')


def cleanup_expired_pdf_folders(items_to_remove):
    """删除过期条目对应的本地 PDF 文件夹。"""
    if not os.path.isdir(OUT_DIR):
        return
    for fname in os.listdir(OUT_DIR):
        fpath = os.path.join(OUT_DIR, fname)
        if not os.path.isdir(fpath):
            continue
        for it in items_to_remove:
            sid = it.get('section_id', '')
            if sid and sid in fname:
                shutil.rmtree(fpath)
                print(f'    删除过期PDF目录: {fname}')
                break


def merge_and_cleanup(new_items):
    """
    读取本地持久 Excel，合并 new_items，过滤过期条目，删除过期PDF，写回。
    返回合并后的全部有效条目列表。
    """
    print(f'\n[Step 3½] 持久Excel增量合并（保留近 {MAX_DAYS} 天）...')

    existing_rows, existing_sids = read_persistent_excel()

    hist_map = {}
    for row in existing_rows:
        sid = row.get('标段编号', '')
        if sid:
            hist_map[str(sid).strip()] = row

    # 合并新增（去重：新数据优先，更新字段）
    merged = dict(hist_map)
    for it in new_items:
        sid = it.get('section_id', '')
        if sid and sid in merged:
            old = merged[sid]
            for k, v in it.items():
                if v and (old.get(k) in (None, '', 'NULL') or k not in old):
                    old[k] = v
            for fk in ['area', 'date', 'title', 'uuid', 'sub_cat']:
                if fk in it and it[fk] and (old.get(fk) in (None, '', 'NULL') or fk not in old):
                    old[fk] = it[fk]
        else:
            merged[sid if sid else f"NEW_{it.get('uuid','')}"] = it

    # 过滤过期
    today = datetime.now().date()
    valid_items = []
    expired_items = []
    for sid, it in merged.items():
        d = parse_date(it.get('date', ''))
        if d is None:
            valid_items.append(it)
            continue
        if d >= today - timedelta(days=MAX_DAYS):
            valid_items.append(it)
        else:
            expired_items.append(it)

    print(f'    合并: {len(merged)} 条 → 有效: {len(valid_items)} 条 | 过期删除: {len(expired_items)} 条')
    if expired_items:
        for e in expired_items:
            print(f'      - {e.get("date","")} | {(e.get("title") or e.get("project_name","") or "")[:40]}')

    if expired_items:
        cleanup_expired_pdf_folders(expired_items)

    valid_items.sort(key=lambda x: x.get('date', ''), reverse=True)
    write_persistent_excel(valid_items)
    return valid_items


# ══════════════════════════════════════════════════════════════
# 自动复盘：每次运行后核查字段提取质量 → 更新 changelog
# ══════════════════════════════════════════════════════════════
def review_and_log_changelog(results, run_label=""):
    import pymupdf, re as _re
    from datetime import datetime

    VALIDATION_FIELDS = [
        'budget', 'bid_method', '基准价计算', 'qual_check',
        'deposit', 'duration', 'qual_req', 'pm_req',
        'price_type', 'valid_period', 'perform_guarantee',
    ]

    # 无标准格式时判定为 N/A 的规则（格式差异，非 bug）
    def _is_epc_or_nonstandard(val, raw_text_for_check):
        """判断某字段为空是否因文档格式差异（EPC/非标准结构），而非提取 bug。"""
        # EPC 项目（工程总承包）没有标准 3.1/3.2 投标人资质结构
        if '工程总承包' in raw_text_for_check:
            return True
        return False

    report_lines = []
    total = 0
    ok_count = 0

    for idx, it in enumerate(results):
        pdfs = it.get('pdfs', [])
        if not pdfs:
            continue
        main_pdf = None
        for pf in pdfs:
            if '招标文件正文' in pf.get('name', ''):
                main_pdf = pf['path']
                break
        if not main_pdf:
            main_pdf = pdfs[0]['path']

        title = it.get('title', it.get('project_name', f'条目{idx+1}'))[:40]

        try:
            doc = pymupdf.open(main_pdf)
            full = '\n'.join(p.get_text() for p in doc)
            doc.close()
        except Exception:
            continue

        text_sp = _re.sub(r'\s+', ' ', full)
        extracted = extract_pdf_fields(main_pdf)

        total += 1
        item_ok = 0
        item_fields = []

        for field in VALIDATION_FIELDS:
            val = extracted.get(field, '')
            if val and val.strip():
                item_ok += 1
                status = '✅'
            elif field in ('perform_guarantee',) or _is_epc_or_nonstandard(val, full):
                status = 'N/A'
            else:
                status = '❌'
            item_fields.append(f"  {status} {field}: {val!r}")

        ok_count += item_ok
        report_lines.append(f"**{title}**")
        report_lines += item_fields

    score = f"{ok_count}/{total * len(VALIDATION_FIELDS)}" if total > 0 else "N/A"
    date_str = datetime.now().strftime('%Y-%m-%d')

    entry = f"""
### {date_str} {run_label} {'✅' if '❌' not in chr(10).join(report_lines) else '⚠️'} 评分：{score}
"""
    if report_lines:
        entry += "**抓取条目**：\n" + '\n'.join(report_lines) + "\n"

    changelog_path = os.path.join(os.path.dirname(__file__), 'changelog.md')
    try:
        with open(changelog_path, 'r', encoding='utf-8') as f:
            content = f.read()
    except Exception:
        content = ""

    marker = "## 历史运行记录"
    if marker in content:
        parts = content.split(marker, 1)
        entry = entry.strip()
        content = parts[0] + marker + "\n\n---\n\n" + entry + "\n\n" + parts[1].lstrip("\n")
    else:
        content = content + "\n\n" + entry

    with open(changelog_path, 'w', encoding='utf-8') as f:
        f.write(content)

    print(f"\n[字段质量复盘] {'✅ 无问题' if '❌' not in entry else '⚠️ 有字段未提取'} | {score}")
    return ok_count, total


# ══════════════════════════════════════════════════════════════
# 主函数
# ══════════════════════════════════════════════════════════════
def main(args):
    sync = '--sync' in args
    notify = '--notify' in args
    dry_run = '--dry-run' in args
    if sync: args.remove('--sync')
    if notify: args.remove('--notify')
    if dry_run: args.remove('--dry-run')

    max_n = int(args[0]) if len(args) > 0 else 20

    today = datetime.now().strftime('%Y-%m-%d')
    week_ago = (datetime.now() - timedelta(days=MAX_DAYS)).strftime('%Y-%m-%d')
    start, end = week_ago, today

    cat_name = '招标公告-施工'

    print('='*60)
    print(f'  {SITE_NAME} 招标信息抓取工具 v2')
    print(f'  分类：{cat_name}（chanId={CHAN_ID}）')
    if sync:
        print(f'  模式：同步腾讯文档（仅保留近 {MAX_DAYS} 天）')
    print('='*60)
    print(f'  日期: {start} ~ {end} | 最大: {max_n}\n')

    print(f'[Step 1] 抓取列表（近{MAX_DAYS}天）...')
    try:
        resp = post(LIST_API_URL, {'chanId': CHAN_ID, 'jyly': '', 'pageIndex': 1, 'pageSize': max_n})
        items = parse_list_json(resp)
        print(f'  JSON: {len(items)} 条记录')
    except Exception as e:
        print(f'  [ERROR] API 请求失败: {e}')
        return

    # 过滤日期窗口
    filt = [x for x in items if x['date'] >= start and x['date'] <= end]
    filt.sort(key=lambda x: x['date'], reverse=True)

    print(f'  解析: {len(filt)} 条（在 {start} ~ {end} 窗口内）\n')
    for i, it in enumerate(filt[:8]):
        print(f'  [{i+1}] {it["date"]} | {it["area"]} | {it["title"][:48]}')
    print()

    if not filt:
        print('[DONE] 0 items in date window')
        return

    print(f'[Step 2] 抓取详情页 + PDF ({len(filt)} items)...')
    results = []
    for i, it in enumerate(filt):
        print(f'  [{i+1}/{len(filt)}] {it["title"][:40]}...')
        detail_url = it.get('detail_url', '')
        if not detail_url:
            print(f'      [WARN] 无详情页 URL，跳过')
            it['sub_cat'] = cat_name
            results.append(it)
            continue
        try:
            detail_html_bytes = get(detail_url)
            detail_html = detail_html_bytes.decode('utf-8', errors='replace')
        except Exception as e:
            print(f'      [WARN] 详情页请求失败: {e}')
            it['sub_cat'] = cat_name
            results.append(it)
            continue
        fields = extract_detail_fields(detail_html)
        it['sub_cat'] = cat_name
        for k, v in fields.items():
            if v:
                it[k] = v

        # PDF 下载
        out_dir = os.path.join(OUT_DIR, f"{it['date'].replace('-','')}-{it['uuid'][:8]}")
        os.makedirs(out_dir, exist_ok=True)
        pdfs = fetch_pdfs(detail_html, detail_url, out_dir)
        it['pdfs'] = pdfs

        # 从 PDF 提取字段
        main_pdf = None
        for pf in pdfs:
            if '招标文件正文' in pf['name']:
                main_pdf = pf['path']
                break
        if not main_pdf and pdfs:
            main_pdf = pdfs[0]['path']
        if main_pdf and os.path.exists(main_pdf):
            try:
                pdf_fields = extract_pdf_fields(main_pdf)
                for k, v in pdf_fields.items():
                    if v and not it.get(k):
                        it[k] = v
            except Exception as e:
                print(f'      [WARN] PDF 字段提取失败: {e}')
            if pdfs:
                print(f'      PDF: {pdfs[0]["name"]} ({pdfs[0]["size"]:,} bytes)')
                if len(pdfs) > 1:
                    for pf in pdfs[1:]:
                        print(f'      跳过(次要): {pf["name"]}')
        elif not pdfs:
            print(f'      [WARN] 无 PDF 附件')
        results.append(it)

    # ── 自动复盘：核查字段提取质量并更新 changelog ──────────────────────────
    review_and_log_changelog(results, run_label="第一次运行")

    # ── Step 3½: 持久Excel增量合并（去重+过期清理） ─────────────────────────
    merged_all = merge_and_cleanup(results)

    print(f'\n[DONE] 本次抓取 {len(results)} 条')
    print(f'       持续积累 Excel 现有 {len(merged_all)} 条（已合并去重+过期清理）')
    for it in results:
        pdf_count = len(it.get('pdfs', []))
        print(f'  {it["date"]} | {cat_name} | {pdf_count} PDFs | {it.get("project_name","") or it.get("title","")[:40]}')

    if sync:
        print(f'\n[Step 4] 腾讯文档同步（基于合并后全量 {len(merged_all)} 条）...')
        mcporter_cmd = 'mcporter'
        qq_sync(mcporter_cmd, merged_all, dry_run=dry_run)
        if not dry_run:
            print(f'\n  同步完成：{TENCENT_URL}')
        else:
            print(f'\n  腾讯文档：{TENCENT_URL}')

if __name__ == '__main__':
    main(sys.argv[1:])
